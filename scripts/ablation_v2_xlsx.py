#!/usr/bin/env python3
"""
Ablation V2 — 466 combinations → Excel (10 sheets)
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = '/Volumes/soulmaten/POTAL/7field_benchmark'
with open(f'{BASE}/ablation_v2_results.json') as f:
    results = json.load(f)

OUTPUT = f'{BASE}/POTAL_Ablation_V2.xlsx'

# ── Styles ──
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='2F5496')
SUBTITLE_FONT = Font(name='Arial', size=12, bold=True)
BOLD_FONT = Font(name='Arial', size=11, bold=True)
NORMAL_FONT = Font(name='Arial', size=11)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Accuracy colors
def acc_fill(pct):
    if pct >= 100: return PatternFill(start_color='C6EFCE', fill_type='solid')
    if pct >= 90:  return PatternFill(start_color='E2EFDA', fill_type='solid')
    if pct >= 70:  return PatternFill(start_color='FFEB9C', fill_type='solid')
    if pct >= 50:  return PatternFill(start_color='FFD9B3', fill_type='solid')
    return PatternFill(start_color='FFC7CE', fill_type='solid')

def impact_fill(val):
    a = abs(val)
    if a > 30: return PatternFill(start_color='FF4444', fill_type='solid')
    if a > 15: return PatternFill(start_color='FFA500', fill_type='solid')
    if a > 5:  return PatternFill(start_color='FFEB9C', fill_type='solid')
    return PatternFill(start_color='C6EFCE', fill_type='solid')

def impact_font(val):
    if abs(val) > 30: return Font(name='Arial', size=11, bold=True, color='FFFFFF')
    return NORMAL_FONT

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def auto_width(ws, ncol, mx=45):
    for c in range(1, ncol + 1):
        ml = max((len(str(cell.value or '')) for cell in ws[get_column_letter(c)]), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(ml + 2, mx)

def write_level_sheet(ws, title, data):
    """Write a level sheet with standard columns"""
    row = 1
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    row = 3

    headers = ['Removed Fields', 'Used Fields', 'Fields', 'Section%', 'Chapter%', 'Heading%', 'HS6%',
               'ΔSection', 'ΔChapter', 'ΔHeading', 'ΔHS6', 'First Fail Step']
    ncol = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, ncol)
    row += 1

    # Sort by HS6 ascending (worst first)
    sorted_data = sorted(data, key=lambda r: r['step_hs6_pct'])

    for r in sorted_data:
        removed = ', '.join(r['removed_fields']) or '(none)'
        used = ', '.join(r['used_fields'])
        ws.cell(row=row, column=1, value=removed).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=used[:55]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=r['level']).font = NORMAL_FONT

        for ci, key in enumerate(['step_section_pct', 'step_chapter_pct', 'step_heading_pct', 'step_hs6_pct'], 4):
            v = r[key]
            cell = ws.cell(row=row, column=ci, value=f"{v}%")
            cell.font = NORMAL_FONT
            cell.fill = acc_fill(v)

        for ci, key in enumerate(['step_section_pct', 'step_chapter_pct', 'step_heading_pct', 'step_hs6_pct'], 8):
            d = r[key] - 100
            cell = ws.cell(row=row, column=ci, value=f"{d:+d}%" if d != 0 else "0%")
            cell.font = NORMAL_FONT
            if d < 0:
                cell.fill = acc_fill(r[key])

        ws.cell(row=row, column=12, value=r.get('first_fail_step', '')).font = NORMAL_FONT

        for c in range(1, ncol + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True)
        row += 1

    auto_width(ws, ncol)
    return row


wb = Workbook()

# ═══════════════ Sheet 1: Dashboard ═══════════════
ws = wb.active
ws.title = 'Dashboard'
ws.sheet_properties.tabColor = '2F5496'

row = 1
ws.cell(row=row, column=1, value='POTAL v3 Pipeline — Systematic Ablation Test (466 Combinations × 50 Products)').font = TITLE_FONT
ws.merge_cells('A1:I1')
row += 2

info = [
    ('Total Combinations', 466),
    ('Total Pipeline Runs', 23300),
    ('Products Tested', '50 (Amazon real seller data)'),
    ('Baseline', '9/9 fields → 100% at all steps'),
    ('⚠️ Ground Truth', 'Self-referential (baseline = POTAL output, no independent verification)'),
]
for label, val in info:
    ws.cell(row=row, column=1, value=label).font = BOLD_FONT
    ws.cell(row=row, column=2, value=str(val)).font = NORMAL_FONT
    row += 1

row += 1
ws.cell(row=row, column=1, value='Level Summary').font = SUBTITLE_FONT
row += 1

lvl_headers = ['Level', 'Fields', 'Combos', 'Avg Section%', 'Avg Chapter%', 'Avg Heading%', 'Avg HS6%', 'Min HS6%', 'Max HS6%', '100% Combos']
for c, h in enumerate(lvl_headers, 1):
    ws.cell(row=row, column=c, value=h)
style_header(ws, row, len(lvl_headers))
row += 1

for level in range(9, 2, -1):
    lr = [r for r in results if r['level'] == level]
    n = len(lr)
    if n == 0: continue
    avgS = round(sum(r['step_section_pct'] for r in lr) / n)
    avgC = round(sum(r['step_chapter_pct'] for r in lr) / n)
    avgH = round(sum(r['step_heading_pct'] for r in lr) / n)
    avg6 = round(sum(r['step_hs6_pct'] for r in lr) / n)
    min6 = min(r['step_hs6_pct'] for r in lr)
    max6 = max(r['step_hs6_pct'] for r in lr)
    perf = sum(1 for r in lr if r['step_hs6_pct'] == 100)

    vals = [f'Level {level}', f'{level} fields', n, f'{avgS}%', f'{avgC}%', f'{avgH}%', f'{avg6}%', f'{min6}%', f'{max6}%', perf]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if c in (4,5,6,7) and isinstance(v, str) and v.endswith('%'):
            cell.fill = acc_fill(int(v.replace('%','')))
    row += 1

row += 1
ws.cell(row=row, column=1, value='Step-by-Step Drop Analysis (average across all combos per level)').font = SUBTITLE_FONT
row += 1

drop_headers = ['Level', 'Avg Section→Chapter Drop', 'Avg Chapter→Heading Drop', 'Avg Heading→HS6 Drop']
for c, h in enumerate(drop_headers, 1):
    ws.cell(row=row, column=c, value=h)
style_header(ws, row, len(drop_headers))
row += 1

for level in range(9, 2, -1):
    lr = [r for r in results if r['level'] == level]
    n = len(lr)
    if n == 0: continue
    avgS = sum(r['step_section_pct'] for r in lr) / n
    avgC = sum(r['step_chapter_pct'] for r in lr) / n
    avgH = sum(r['step_heading_pct'] for r in lr) / n
    avg6 = sum(r['step_hs6_pct'] for r in lr) / n
    d1 = round(avgC - avgS, 1)
    d2 = round(avgH - avgC, 1)
    d3 = round(avg6 - avgH, 1)
    vals = [f'Level {level}', f'{d1:+.1f}%', f'{d2:+.1f}%', f'{d3:+.1f}%']
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
    row += 1

row += 1
ws.cell(row=row, column=1, value='Key Findings').font = SUBTITLE_FONT
row += 1
findings = [
    '1. material is CRITICAL — removing it alone drops HS6 from 100% to 40% (-60%)',
    '2. category is CRITICAL — removing it alone drops Heading from 100% to 60% (-40%)',
    '3. "Magic 3" = product_name + material + category → achieves 98% HS6 with only 3 fields',
    '4. origin_country, processing, composition, weight_spec, price have ≤2% individual impact',
    '5. At Level 3 (3 fields): best combo is name+material+category (98% HS6), worst is 0%',
    '6. At Level 5 (5 fields): 5 combos still achieve 100% — all include material+category',
    '7. Section step is most resilient (avg 49% even at Level 3), HS6 degrades fastest',
    '8. ⚠️ All accuracy numbers are self-referential (vs baseline, not independent ground truth)',
]
for f in findings:
    ws.cell(row=row, column=1, value=f).font = NORMAL_FONT
    row += 1

auto_width(ws, 10, 55)

# ═══════════════ Sheets 2-7: Level sheets ═══════════════
level_colors = ['808080', 'A0A0A0', 'C0C0C0', 'FFD9B3', 'FFA500', 'FF4444']
for i, level in enumerate(range(8, 2, -1)):
    lr = [r for r in results if r['level'] == level]
    remove_count = 9 - level
    sheet_name = f'Level {level} (Remove {remove_count})'
    ws_l = wb.create_sheet(sheet_name)
    ws_l.sheet_properties.tabColor = level_colors[i]
    write_level_sheet(ws_l, f'Level {level}: Remove {remove_count} Field{"s" if remove_count > 1 else ""} ({len(lr)} combinations)', lr)

# ═══════════════ Sheet 8: Field Importance Matrix ═══════════════
ws8 = wb.create_sheet('Field Importance Matrix')
ws8.sheet_properties.tabColor = '7030A0'

row = 1
ws8.cell(row=row, column=1, value='Field × Step Importance Matrix (466 combinations)').font = TITLE_FONT
ws8.merge_cells('A1:L1')
row += 2

ALL_FIELDS = ['product_name', 'material', 'origin_country', 'category', 'description', 'processing', 'composition', 'weight_spec', 'price']

m_headers = ['Field', 'Incl. Avg S%', 'Incl. Avg Ch%', 'Incl. Avg H%', 'Incl. Avg HS6%',
             'Excl. Avg S%', 'Excl. Avg Ch%', 'Excl. Avg H%', 'Excl. Avg HS6%',
             'S Impact', 'Ch Impact', 'H Impact', 'HS6 Impact']
for c, h in enumerate(m_headers, 1):
    ws8.cell(row=row, column=c, value=h)
style_header(ws8, row, len(m_headers))
row += 1

field_impacts = []
for field in ALL_FIELDS:
    incl = [r for r in results if field in r['used_fields']]
    excl = [r for r in results if field not in r['used_fields']]

    def avg(lst, key):
        return round(sum(r[key] for r in lst) / len(lst), 1) if lst else 0

    iS = avg(incl, 'step_section_pct')
    iC = avg(incl, 'step_chapter_pct')
    iH = avg(incl, 'step_heading_pct')
    i6 = avg(incl, 'step_hs6_pct')
    eS = avg(excl, 'step_section_pct')
    eC = avg(excl, 'step_chapter_pct')
    eH = avg(excl, 'step_heading_pct')
    e6 = avg(excl, 'step_hs6_pct')
    dS = round(iS - eS, 1)
    dC = round(iC - eC, 1)
    dH = round(iH - eH, 1)
    d6 = round(i6 - e6, 1)

    field_impacts.append((field, d6))

    vals = [field, f'{iS}%', f'{iC}%', f'{iH}%', f'{i6}%',
            f'{eS}%', f'{eC}%', f'{eH}%', f'{e6}%',
            f'{dS:+.1f}', f'{dC:+.1f}', f'{dH:+.1f}', f'{d6:+.1f}']
    for c, v in enumerate(vals, 1):
        cell = ws8.cell(row=row, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if c >= 10:
            try:
                fv = float(v)
                cell.fill = impact_fill(fv)
                cell.font = impact_font(fv)
            except:
                pass
    row += 1

# Sort and show ranking
row += 1
ws8.cell(row=row, column=1, value='Field Ranking by HS6 Impact (highest → lowest)').font = SUBTITLE_FONT
row += 1
field_impacts.sort(key=lambda x: -x[1])
for i, (field, impact) in enumerate(field_impacts, 1):
    grade = 'CRITICAL' if impact > 15 else ('HIGH' if impact > 5 else ('LOW' if impact > 1 else 'NONE'))
    ws8.cell(row=row, column=1, value=f'{i}. {field}').font = BOLD_FONT
    ws8.cell(row=row, column=2, value=f'+{impact:.1f}% HS6').font = NORMAL_FONT
    ws8.cell(row=row, column=3, value=grade).font = NORMAL_FONT
    cell_grade = ws8.cell(row=row, column=3)
    colors = {'CRITICAL': 'FF4444', 'HIGH': 'FFA500', 'LOW': 'FFEB9C', 'NONE': 'C6EFCE'}
    cell_grade.fill = PatternFill(start_color=colors.get(grade, 'C6EFCE'), fill_type='solid')
    if grade == 'CRITICAL':
        cell_grade.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    row += 1

auto_width(ws8, len(m_headers), 20)

# ═══════════════ Sheet 9: Step Failure Heatmap ═══════════════
ws9 = wb.create_sheet('Step Failure Heatmap')
ws9.sheet_properties.tabColor = 'FFA500'

row = 1
ws9.cell(row=row, column=1, value='Step Failure Heatmap — Average Accuracy by Level × Step').font = TITLE_FONT
ws9.merge_cells('A1:E1')
row += 2

hm_headers = ['Level', 'Step 2-1/2-2 (Section)', 'Step 2-3/2-4 (Chapter)', 'Step 3 (Heading)', 'Step 4 (HS6)']
for c, h in enumerate(hm_headers, 1):
    ws9.cell(row=row, column=c, value=h)
style_header(ws9, row, len(hm_headers))
row += 1

for level in range(9, 2, -1):
    lr = [r for r in results if r['level'] == level]
    n = len(lr)
    if n == 0: continue
    avgS = round(sum(r['step_section_pct'] for r in lr) / n)
    avgC = round(sum(r['step_chapter_pct'] for r in lr) / n)
    avgH = round(sum(r['step_heading_pct'] for r in lr) / n)
    avg6 = round(sum(r['step_hs6_pct'] for r in lr) / n)

    ws9.cell(row=row, column=1, value=f'Level {level} ({level}F, {n} combos)').font = BOLD_FONT
    ws9.cell(row=row, column=1).border = THIN_BORDER
    for ci, v in enumerate([avgS, avgC, avgH, avg6], 2):
        cell = ws9.cell(row=row, column=ci, value=f'{v}%')
        cell.font = BOLD_FONT
        cell.fill = acc_fill(v)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

row += 2
ws9.cell(row=row, column=1, value='Reading Guide:').font = SUBTITLE_FONT
row += 1
guides = [
    'Green (100%) = no degradation at this step',
    'Light green (90-99%) = minor degradation',
    'Yellow (70-89%) = significant degradation — this step starts breaking',
    'Orange (50-69%) = major degradation',
    'Red (<50%) = step is unreliable at this field count',
    '',
    'Pattern: Section degrades slowest (material-based), HS6 degrades fastest (needs more detail)',
    'Biggest cliff: Level 7→6 at Heading step, Level 5→4 at all steps',
]
for g in guides:
    ws9.cell(row=row, column=1, value=g).font = NORMAL_FONT
    row += 1

auto_width(ws9, 5, 35)

# ═══════════════ Sheet 10: Raw Data ═══════════════
ws10 = wb.create_sheet('Raw Data')
ws10.sheet_properties.tabColor = '808080'

raw_headers = ['combo_id', 'level', 'removed_fields', 'used_fields', 'field_count',
               'section%', 'chapter%', 'heading%', 'hs6%', 'first_fail_step']
for c, h in enumerate(raw_headers, 1):
    ws10.cell(row=1, column=c, value=h)
style_header(ws10, 1, len(raw_headers))

for i, r in enumerate(results):
    row = i + 2
    ws10.cell(row=row, column=1, value=r['combo_id']).font = NORMAL_FONT
    ws10.cell(row=row, column=2, value=r['level']).font = NORMAL_FONT
    ws10.cell(row=row, column=3, value=', '.join(r['removed_fields'])).font = NORMAL_FONT
    ws10.cell(row=row, column=4, value=', '.join(r['used_fields'])[:55]).font = NORMAL_FONT
    ws10.cell(row=row, column=5, value=r['level']).font = NORMAL_FONT
    for ci, key in enumerate(['step_section_pct', 'step_chapter_pct', 'step_heading_pct', 'step_hs6_pct'], 6):
        cell = ws10.cell(row=row, column=ci, value=r[key])
        cell.font = NORMAL_FONT
        cell.fill = acc_fill(r[key])
    ws10.cell(row=row, column=10, value=r.get('first_fail_step', '')).font = NORMAL_FONT
    for c in range(1, len(raw_headers) + 1):
        ws10.cell(row=row, column=c).border = THIN_BORDER

auto_width(ws10, len(raw_headers), 55)

# ── Save ──
wb.save(OUTPUT)
print(f'✅ Excel saved: {OUTPUT}')
print(f'   10 sheets: Dashboard, Level 8-3 (6 sheets), Field Importance Matrix, Step Failure Heatmap, Raw Data')
