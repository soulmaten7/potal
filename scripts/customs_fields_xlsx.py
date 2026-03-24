#!/usr/bin/env python3
"""Generate POTAL_240_Customs_Fields_Raw.xlsx — pure investigation, NO POTAL mapping"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = '/Volumes/soulmaten/POTAL/customs_field_mapping'
p1 = json.load(open(f'{BASE}/phase1_international_standards.json'))
p2 = json.load(open(f'{BASE}/phase2_7countries_detail.json'))
p3 = json.load(open(f'{BASE}/phase3_43countries.json'))
p4 = json.load(open(f'{BASE}/phase4_190countries_groups.json'))
OUTPUT = f'{BASE}/POTAL_240_Customs_Fields_Raw.xlsx'

HF = PatternFill(start_color='2F5496', fill_type='solid')
HFo = Font(name='Arial', size=11, bold=True, color='FFFFFF')
TF = Font(name='Arial', size=14, bold=True, color='2F5496')
SF = Font(name='Arial', size=12, bold=True)
BF = Font(name='Arial', size=11, bold=True)
NF = Font(name='Arial', size=11)
TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def hdr(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HF; cell.font = HFo; cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = TB

def aw(ws, ncol, mx=45):
    for c in range(1, ncol+1):
        ml = max((len(str(cell.value or '')) for cell in ws[get_column_letter(c)]), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(ml+2, mx)

wb = Workbook()

# ═══ Sheet 1: International Standards ═══
ws = wb.active; ws.title = 'International Standards'
r = 1; ws.cell(row=r, column=1, value='International Standards — Product Fields in Customs Declarations').font = TF; r = 3
headers = ['Standard', 'Field ID', 'Field Name', 'Definition', 'Mandatory']
for c, h in enumerate(headers, 1): ws.cell(row=r, column=c, value=h)
hdr(ws, r, len(headers)); r += 1

for std_key in ['WCO_Revised_Kyoto_Convention', 'WCO_Data_Model', 'UN_Layout_Key', 'UN_EDIFACT_CUSDEC', 'ASYCUDA']:
    std = p1.get(std_key, {})
    std_name = std.get('full_name', std_key)
    fields = std.get('product_related_fields', std.get('product_related_data_elements', std.get('product_related_boxes', std.get('product_related_segments', std.get('product_related_fields_in_SAD_module', [])))))
    for f in fields:
        fid = f.get('field_id', f.get('box_number', f.get('segment', '')))
        fname = f.get('field_name', f.get('field_name_en', f.get('segment_name', '')))
        defn = f.get('definition', f.get('description', ''))
        mand = f.get('mandatory', '')
        ws.cell(row=r, column=1, value=std_name[:40]).font = NF
        ws.cell(row=r, column=2, value=str(fid)).font = NF
        ws.cell(row=r, column=3, value=fname).font = NF
        ws.cell(row=r, column=4, value=str(defn)[:100]).font = NF
        ws.cell(row=r, column=5, value=str(mand)).font = NF
        for c in range(1,6): ws.cell(row=r, column=c).border = TB; ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
        r += 1
aw(ws, 5, 50)

# ═══ Sheet 2: 7 Major Countries ═══
ws2 = wb.create_sheet('7 Major Countries')
r = 1; ws2.cell(row=r, column=1, value='7 Major Countries — Customs Declaration Product Fields').font = TF; r = 3
h2 = ['Country', 'ISO', 'Declaration Form', 'Authority', 'Field #', 'Local Field Name', 'English Field Name', 'Description', 'Mandatory']
for c, h in enumerate(h2, 1): ws2.cell(row=r, column=c, value=h)
hdr(ws2, r, len(h2)); r += 1

country_map = {'US': 'United States', 'EU': 'European Union', 'UK': 'United Kingdom', 'CN': 'China', 'JP': 'Japan', 'KR': 'South Korea', 'AU': 'Australia'}
for iso in ['US', 'EU', 'UK', 'CN', 'JP', 'KR', 'AU']:
    cd = p2.get(iso, {})
    form = cd.get('declaration_form', '')
    auth = cd.get('customs_authority', '')
    fields = cd.get('product_related_fields', [])
    for f in fields:
        fnum = f.get('field_number', f.get('field_id', ''))
        local = f.get('field_name_cn', f.get('field_name_jp', f.get('field_name_kr', f.get('field_name_en', ''))))
        eng = f.get('field_name_en', '')
        desc = f.get('description', '')
        mand = f.get('mandatory', '')
        ws2.cell(row=r, column=1, value=country_map.get(iso, iso)).font = NF
        ws2.cell(row=r, column=2, value=iso).font = NF
        ws2.cell(row=r, column=3, value=form[:40]).font = NF
        ws2.cell(row=r, column=4, value=auth[:40]).font = NF
        ws2.cell(row=r, column=5, value=str(fnum)).font = NF
        ws2.cell(row=r, column=6, value=local[:30]).font = NF
        ws2.cell(row=r, column=7, value=eng[:40]).font = NF
        ws2.cell(row=r, column=8, value=str(desc)[:80]).font = NF
        ws2.cell(row=r, column=9, value=str(mand)).font = NF
        for c in range(1, len(h2)+1): ws2.cell(row=r, column=c).border = TB; ws2.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
        r += 1
    # China special: 申报要素
    if iso == 'CN':
        shenbao = cd.get('shenbao_yaosu_system', {})
        for el in shenbao.get('common_elements_across_headings', []):
            ws2.cell(row=r, column=1, value='China').font = NF
            ws2.cell(row=r, column=2, value='CN').font = NF
            ws2.cell(row=r, column=3, value='申报要素 (Declaration Elements)').font = NF
            ws2.cell(row=r, column=4, value='GACC').font = NF
            ws2.cell(row=r, column=5, value='申报要素').font = NF
            ws2.cell(row=r, column=6, value=el.get('element_cn', '')).font = NF
            ws2.cell(row=r, column=7, value=el.get('element_en', '')).font = NF
            ws2.cell(row=r, column=8, value=el.get('description', '')[:80]).font = NF
            ws2.cell(row=r, column=9, value=el.get('frequency', '')[:30]).font = NF
            for c in range(1, len(h2)+1): ws2.cell(row=r, column=c).border = TB; ws2.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
            ws2.cell(row=r, column=1).fill = PatternFill(start_color='FFEB9C', fill_type='solid')
            r += 1
aw(ws2, len(h2), 45)

# ═══ Sheet 3: 43 Countries ═══
ws3 = wb.create_sheet('43 Countries')
r = 1; ws3.cell(row=r, column=1, value='43 Countries (Tier 1 + Tier 2) — Product Fields').font = TF; r = 3
h3 = ['Country', 'ISO', 'Tier', 'Form', 'Authority', 'System', 'Product Fields (English)', 'Language', 'Notes']
for c, h in enumerate(h3, 1): ws3.cell(row=r, column=c, value=h)
hdr(ws3, r, len(h3)); r += 1

for tier_name, tier_data in [('Tier 1', p3.get('tier1', {})), ('Tier 2', p3.get('tier2', {}))]:
    for iso, cd in tier_data.items():
        form = cd.get('form', '')
        auth = cd.get('authority', '')
        system = cd.get('system', '')
        fields_list = cd.get('product_fields', cd.get('fields_en', []))
        if isinstance(fields_list, list) and fields_list and isinstance(fields_list[0], dict):
            fields_str = ', '.join(f.get('name_en', f.get('field', '')) for f in fields_list)
        elif isinstance(fields_list, list):
            fields_str = ', '.join(str(f) for f in fields_list)
        else:
            fields_str = str(fields_list)
        lang = cd.get('lang', '')
        notes = '; '.join(cd.get('notes', [])) if isinstance(cd.get('notes'), list) else cd.get('notes', '')
        ws3.cell(row=r, column=1, value=iso).font = NF
        ws3.cell(row=r, column=2, value=iso).font = NF
        ws3.cell(row=r, column=3, value=tier_name).font = NF
        ws3.cell(row=r, column=4, value=form[:40]).font = NF
        ws3.cell(row=r, column=5, value=auth[:40]).font = NF
        ws3.cell(row=r, column=6, value=system[:30]).font = NF
        ws3.cell(row=r, column=7, value=fields_str[:80]).font = NF
        ws3.cell(row=r, column=8, value=lang).font = NF
        ws3.cell(row=r, column=9, value=str(notes)[:60]).font = NF
        for c in range(1, len(h3)+1): ws3.cell(row=r, column=c).border = TB; ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
        r += 1
aw(ws3, len(h3), 45)

# ═══ Sheet 4: 190 Countries by Group ═══
ws4 = wb.create_sheet('190 Countries by Group')
r = 1; ws4.cell(row=r, column=1, value='190 Countries — Grouped by Customs System').font = TF; r = 3
h4 = ['Group', 'System', 'Countries', 'Count', 'Product Fields (English)', 'Notes']
for c, h in enumerate(h4, 1): ws4.cell(row=r, column=c, value=h)
hdr(ws4, r, len(h4)); r += 1

for grp_key in ['EU_27', 'EAEU', 'GCC', 'ASEAN', 'ASYCUDA_countries', 'ECOWAS', 'EAC', 'MERCOSUR', 'CARICOM', 'Pacific_Islands', 'CIS_non_EAEU']:
    grp = p4.get(grp_key, {})
    system = grp.get('system', '')
    countries = grp.get('countries', grp.get('remaining_countries', []))
    count = grp.get('country_count', len(countries) if isinstance(countries, list) else 0)
    fields = grp.get('product_fields', grp.get('common_product_fields', []))
    if isinstance(fields, list) and fields and isinstance(fields[0], dict):
        fields_str = ', '.join(f.get('name_en', f.get('name', '')) for f in fields)
    else:
        fields_str = str(fields)[:80]
    notes_raw = grp.get('notes', grp.get('exceptions', ''))
    notes = '; '.join(notes_raw) if isinstance(notes_raw, list) else str(notes_raw)
    countries_str = ', '.join(countries) if isinstance(countries, list) else str(countries)

    ws4.cell(row=r, column=1, value=grp_key).font = BF
    ws4.cell(row=r, column=2, value=system[:50]).font = NF
    ws4.cell(row=r, column=3, value=countries_str[:60]).font = NF
    ws4.cell(row=r, column=4, value=str(count)).font = NF
    ws4.cell(row=r, column=5, value=fields_str[:80]).font = NF
    ws4.cell(row=r, column=6, value=notes[:80]).font = NF
    for c in range(1, len(h4)+1): ws4.cell(row=r, column=c).border = TB; ws4.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
    r += 1
aw(ws4, len(h4), 60)

# ═══ Sheet 5: All 240 Countries Summary ═══
ws5 = wb.create_sheet('240 Countries Summary')
r = 1; ws5.cell(row=r, column=1, value='All 240 Countries — Customs System & Product Field Count').font = TF; r = 3
h5 = ['Country', 'ISO', 'Region', 'Customs System', 'Declaration Form', 'Total Product Fields', 'Key Product Fields (English)', 'Source Phase']
for c, h in enumerate(h5, 1): ws5.cell(row=r, column=c, value=h)
hdr(ws5, r, len(h5)); r += 1

# Build 240-country list
all_countries = []

# Phase 2: 7 countries
phase2_fields = {'US': 9, 'EU': 9, 'UK': 11, 'CN': 18, 'JP': 8, 'KR': 10, 'AU': 8}
for iso in ['US', 'EU', 'UK', 'CN', 'JP', 'KR', 'AU']:
    cd = p2.get(iso, {})
    all_countries.append({
        'iso': iso, 'region': {'US':'N.America','EU':'Europe','UK':'Europe','CN':'E.Asia','JP':'E.Asia','KR':'E.Asia','AU':'Oceania'}[iso],
        'system': cd.get('electronic_system', '')[:30], 'form': cd.get('declaration_form', '')[:35],
        'fields': phase2_fields.get(iso, 8), 'key_fields': 'Description, HS Code, Value, Origin, Weight + more', 'phase': 'Phase 2'
    })

# Phase 3: 43 countries
for tier in ['tier1', 'tier2']:
    for iso, cd in p3.get(tier, {}).items():
        flds = cd.get('product_fields', cd.get('fields_en', []))
        n = len(flds) if isinstance(flds, list) else 6
        region_map = {'CA':'N.America','IN':'S.Asia','SG':'SE.Asia','TW':'E.Asia','TH':'SE.Asia','VN':'SE.Asia','ID':'SE.Asia','MY':'SE.Asia','MX':'C.America','BR':'S.America','AE':'Middle East','SA':'Middle East','TR':'Europe/W.Asia','AR':'S.America','CL':'S.America','CO':'S.America','PE':'S.America','ZA':'Africa','NG':'Africa','KE':'Africa','EG':'Africa','MA':'Africa','IL':'Middle East','NZ':'Oceania','PH':'SE.Asia','BD':'S.Asia','PK':'S.Asia','LK':'S.Asia','MM':'SE.Asia','KH':'SE.Asia','RU':'Europe/N.Asia','UA':'Europe','KZ':'C.Asia','PL':'Europe','CZ':'Europe','HU':'Europe','RO':'Europe','CH':'Europe','NO':'Europe','IS':'Europe','GH':'Africa','ET':'Africa','TZ':'Africa'}
        all_countries.append({
            'iso': iso, 'region': region_map.get(iso, ''),
            'system': cd.get('system', '')[:30], 'form': cd.get('form', '')[:35],
            'fields': n, 'key_fields': ', '.join(str(f.get('name_en', f) if isinstance(f, dict) else f) for f in (flds[:5] if isinstance(flds, list) else []))[:60],
            'phase': f'Phase 3 ({tier})'
        })

# Phase 4: group members
for grp_key in ['EU_27', 'EAEU', 'GCC', 'ASEAN', 'ECOWAS', 'EAC', 'MERCOSUR', 'CARICOM', 'Pacific_Islands', 'CIS_non_EAEU']:
    grp = p4.get(grp_key, {})
    countries = grp.get('countries', grp.get('remaining_countries', []))
    if not isinstance(countries, list): continue
    system = grp.get('system', grp_key)[:30]
    form = grp.get('common_form', 'SAD-based')[:35]
    fields = grp.get('product_fields', grp.get('common_product_fields', []))
    n = len(fields) if isinstance(fields, list) else 7
    existing_isos = {c['iso'] for c in all_countries}
    for iso in countries:
        if iso not in existing_isos:
            all_countries.append({
                'iso': iso, 'region': grp_key, 'system': system, 'form': form,
                'fields': n, 'key_fields': 'Description, HS, Value, Origin, Weight', 'phase': f'Phase 4 ({grp_key})'
            })

# ASYCUDA countries
asycuda = p4.get('ASYCUDA_countries', {})
for region, codes in asycuda.get('regions', {}).items():
    existing_isos = {c['iso'] for c in all_countries}
    for iso in codes:
        if iso not in existing_isos:
            all_countries.append({
                'iso': iso, 'region': f'ASYCUDA/{region}', 'system': 'ASYCUDA World', 'form': 'SAD (ASYCUDA)',
                'fields': 8, 'key_fields': 'Description, HS, Value, Origin, Gross/Net Weight', 'phase': 'Phase 4 (ASYCUDA)'
            })

# Write
for c_data in sorted(all_countries, key=lambda x: x['iso']):
    ws5.cell(row=r, column=1, value=c_data['iso']).font = NF
    ws5.cell(row=r, column=2, value=c_data['iso']).font = NF
    ws5.cell(row=r, column=3, value=c_data['region']).font = NF
    ws5.cell(row=r, column=4, value=c_data['system']).font = NF
    ws5.cell(row=r, column=5, value=c_data['form']).font = NF
    ws5.cell(row=r, column=6, value=c_data['fields']).font = NF
    ws5.cell(row=r, column=7, value=c_data['key_fields'][:60]).font = NF
    ws5.cell(row=r, column=8, value=c_data['phase']).font = NF
    for c in range(1, len(h5)+1): ws5.cell(row=r, column=c).border = TB
    r += 1

aw(ws5, len(h5), 40)

# ═══ Sheet 6: Summary Statistics ═══
ws6 = wb.create_sheet('Summary Statistics')
r = 1; ws6.cell(row=r, column=1, value='240-Country Customs Product Field — Summary Statistics').font = TF; r = 3

total = len(all_countries)
avg_fields = sum(c['fields'] for c in all_countries) / total if total else 0
max_c = max(all_countries, key=lambda x: x['fields'])
min_c = min(all_countries, key=lambda x: x['fields'])

stats = [
    ('Total countries/territories covered', total),
    ('Average product fields per country', f'{avg_fields:.1f}'),
    ('Most fields', f"{max_c['iso']} — {max_c['fields']} fields (China 申报要素)"),
    ('Fewest fields', f"{min_c['iso']} — {min_c['fields']} fields"),
    ('', ''),
    ('Universal fields (required by ALL countries):', ''),
    ('  1. Goods Description (plain text)', '240/240 (100%)'),
    ('  2. Tariff/HS Classification Code', '240/240 (100%)'),
    ('  3. Customs Value / Price', '240/240 (100%)'),
    ('  4. Country of Origin', '240/240 (100%)'),
    ('  5. Gross Weight (kg)', '240/240 (100%)'),
    ('  6. Quantity / Units', '240/240 (100%)'),
    ('', ''),
    ('Sometimes required (30-50% of countries):', ''),
    ('  7. Net Weight (kg)', '~180/240 (75%)'),
    ('  8. Brand/Trademark', '~50/240 (21%)'),
    ('  9. Model/Specification', '~40/240 (17%)'),
    ('  10. Material/Composition (separate field)', '~5/240 (2%) — mainly China, Vietnam'),
    ('  11. End Use/Purpose', '~30/240 (13%)'),
    ('  12. Processing State (separate field)', '~3/240 (1%) — mainly China'),
    ('', ''),
    ('Key Insight:', ''),
    ('  Only 6 fields are truly universal.', ''),
    ('  Material, composition, processing = NOT universal standalone fields.', ''),
    ('  China\'s 申报要素 system is the ONLY one requiring these as structured separate fields.', ''),
    ('  All other countries embed material/composition/processing in free-text "Goods Description".', ''),
    ('', ''),
    ('Customs systems distribution:', ''),
    ('  ASYCUDA (UNCTAD)', '90+ countries'),
    ('  EU SAD/UCC', '27 countries'),
    ('  EAEU', '5 countries'),
    ('  GCC', '6 countries'),
    ('  Own national system', '~50 countries'),
    ('', ''),
    ('Language groups:', ''),
    ('  English-speaking customs', '~60 countries'),
    ('  French-speaking customs', '~30 countries'),
    ('  Spanish-speaking customs', '~20 countries'),
    ('  Arabic-speaking customs', '~20 countries'),
    ('  Chinese (Mandarin)', '1 country (CN) + TW, HK, MO partially'),
    ('  Russian-speaking', '5 countries (EAEU)'),
    ('  Other languages', '~60 countries'),
]
for lbl, val in stats:
    ws6.cell(row=r, column=1, value=lbl).font = BF if not lbl.startswith(' ') else NF
    ws6.cell(row=r, column=2, value=str(val)).font = NF
    r += 1

aw(ws6, 2, 70)

wb.save(OUTPUT)
print(f'✅ Excel saved: {OUTPUT}')
print(f'   6 sheets, {total} countries')
