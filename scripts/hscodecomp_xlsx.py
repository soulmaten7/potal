#!/usr/bin/env python3
"""HSCodeComp 632 benchmark → add sheets to POTAL_Ablation_V2.xlsx"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = '/Volumes/soulmaten/POTAL/7field_benchmark'
results = json.load(open(f'{BASE}/hscodecomp_results.json'))
errors_list = json.load(open(f'{BASE}/hscodecomp_errors.json'))
fixes = json.load(open(f'{BASE}/hscodecomp_fixes.json'))

wb = load_workbook(f'{BASE}/POTAL_Ablation_V2.xlsx')
HF = PatternFill(start_color='2F5496', fill_type='solid')
HFo = Font(name='Arial', size=11, bold=True, color='FFFFFF')
TF = Font(name='Arial', size=14, bold=True, color='2F5496')
SF = Font(name='Arial', size=12, bold=True)
BF = Font(name='Arial', size=11, bold=True)
NF = Font(name='Arial', size=11)
TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
GF = PatternFill(start_color='C6EFCE', fill_type='solid')
RF = PatternFill(start_color='FFC7CE', fill_type='solid')
YF = PatternFill(start_color='FFEB9C', fill_type='solid')
OF = PatternFill(start_color='FFD9B3', fill_type='solid')

def hdr(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HF; cell.font = HFo; cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = TB

def aw(ws, ncol, mx=45):
    for c in range(1, ncol+1):
        ml = max((len(str(cell.value or '')) for cell in ws[get_column_letter(c)]), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(ml+2, mx)

N = 632
ch_ok = fixes['chapter_correct']
h_ok = fixes['heading_correct']
h6_ok = fixes['hs6_correct']

# ═══ Sheet: HSCodeComp Dashboard ═══
ws = wb.create_sheet('HSCodeComp Dashboard')
ws.sheet_properties.tabColor = '00B050'
r = 1
ws.cell(row=r, column=1, value='HSCodeComp 632 Benchmark — Independent Ground Truth').font = TF
ws.merge_cells('A1:H1'); r = 3

for lbl, val in [
    ('Source', 'HuggingFace AIDC-AI/HSCodeComp (public benchmark)'),
    ('Items', '632 (all 10-digit US HTS codes)'),
    ('Ground Truth', 'Confirmed HS codes from AliExpress product data'),
    ('Data Type', 'AliExpress seller listings (Chinese cross-border e-commerce)'),
    ('Available Fields', 'product_name (100%), category (100%), origin (97%), price (91%), weight (87%), material (57%)'),
    ('Missing Fields', 'description (0%), processing (0%), composition (0%)'),
]:
    ws.cell(row=r, column=1, value=lbl).font = BF
    ws.cell(row=r, column=2, value=val).font = NF; r += 1

r += 1; ws.cell(row=r, column=1, value='Overall Accuracy (Ground Truth)').font = SF; r += 1
acc_headers = ['', 'Chapter (2-digit)', 'Heading (4-digit)', 'HS6 (6-digit)']
for c, h in enumerate(acc_headers, 1): ws.cell(row=r, column=c, value=h)
hdr(ws, r, 4); r += 1
for label, vals in [
    ('Correct', [ch_ok, h_ok, h6_ok]),
    ('Total', [N, N, N]),
    ('Accuracy', [f'{ch_ok/N*100:.1f}%', f'{h_ok/N*100:.1f}%', f'{h6_ok/N*100:.1f}%']),
]:
    ws.cell(row=r, column=1, value=label).font = BF
    for c, v in enumerate(vals, 2):
        cell = ws.cell(row=r, column=c, value=v); cell.font = BF; cell.border = TB
        if isinstance(v, str) and '%' in v:
            pct = float(v.replace('%',''))
            if pct >= 50: cell.fill = YF
            elif pct >= 20: cell.fill = OF
            else: cell.fill = RF
    r += 1

r += 1; ws.cell(row=r, column=1, value='Competitor Comparison').font = SF; r += 1
comp_headers = ['Provider', 'HS6 Accuracy', 'Benchmark', 'Method', 'Cost']
for c, h in enumerate(comp_headers, 1): ws.cell(row=r, column=c, value=h)
hdr(ws, r, len(comp_headers)); r += 1
comps = [
    ('POTAL v3', f'{h6_ok/N*100:.1f}%', 'HSCodeComp 632 (public)', 'Code-only pipeline (0 AI calls)', '$0'),
    ('Tarifflo', '89%', '103 items (non-public, self-reported)', 'AI + rules', '$?'),
    ('Avalara', '80%', 'Tarifflo paper test', 'AI (GPT)', '$?'),
    ('Zonos', '44%', 'Tarifflo paper test', 'AI', '$?'),
    ('WCO BACUDA', '13%', 'arXiv paper', 'ML model', '$?'),
    ('AI Best (arXiv)', '46.8%', 'HSCodeComp 632', 'GPT-4 / LLM', '~$0.02/item'),
]
for comp in comps:
    for c, v in enumerate(comp, 1):
        ws.cell(row=r, column=c, value=v).font = NF; ws.cell(row=r, column=c).border = TB
    r += 1

r += 1; ws.cell(row=r, column=1, value='Error Type Distribution').font = SF; r += 1
et_headers = ['Error Type', 'Count', '%', 'Code Fix?']
for c, h in enumerate(et_headers, 1): ws.cell(row=r, column=c, value=h)
hdr(ws, r, len(et_headers)); r += 1
for etype, cnt in sorted(fixes['by_type'].items(), key=lambda x: -x[1]):
    fix = 'YES' if etype not in ('FIELD_DEPENDENT','HS_VERSION_MISMATCH','DATA_AMBIGUOUS') else 'No'
    ws.cell(row=r, column=1, value=etype).font = NF
    ws.cell(row=r, column=2, value=cnt).font = NF
    ws.cell(row=r, column=3, value=f'{cnt/len(errors_list)*100:.1f}%').font = NF
    ws.cell(row=r, column=4, value=fix).font = NF
    if fix == 'YES': ws.cell(row=r, column=4).fill = RF
    for c in range(1,5): ws.cell(row=r, column=c).border = TB
    r += 1

r += 1; ws.cell(row=r, column=1, value='Error by Step').font = SF; r += 1
for step, cnt in sorted(fixes['by_step'].items(), key=lambda x: -x[1]):
    ws.cell(row=r, column=1, value=step).font = NF
    ws.cell(row=r, column=2, value=cnt).font = NF
    ws.cell(row=r, column=3, value=f'{cnt/len(errors_list)*100:.1f}%').font = NF
    for c in range(1,4): ws.cell(row=r, column=c).border = TB
    r += 1

r += 1; ws.cell(row=r, column=1, value='Comparison: Amazon 50 vs HSCodeComp 632').font = SF; r += 1
cmp_h = ['Metric', 'Amazon 50 (self-ref)', 'HSCodeComp 632 (GT)', 'Notes']
for c, h in enumerate(cmp_h, 1): ws.cell(row=r, column=c, value=h)
hdr(ws, r, 4); r += 1
cmps = [
    ('HS6 Accuracy', '100% (50/50)', f'{h6_ok/N*100:.1f}% ({h6_ok}/632)', 'Amazon = self-referential, HSCodeComp = independent GT'),
    ('Chapter Accuracy', '100%', f'{ch_ok/N*100:.1f}%', 'Chapter is where most errors occur'),
    ('Code Bugs Found', '0', f'{fixes["code_fix_needed"]}', 'KEYWORD_MISSING = needs synonym expansion'),
    ('Data Format', 'Seller 9-field', 'AliExpress listing', 'Different category taxonomy'),
    ('Field Count', '9/9', '~4-5/9 avg', 'HSCodeComp missing description/processing/composition'),
    ('Product Types', '10 consumer categories', '40+ categories (jewelry-heavy)', 'HSCodeComp skewed toward jewelry/electronics'),
]
for cmp in cmps:
    for c, v in enumerate(cmp, 1):
        ws.cell(row=r, column=c, value=v).font = NF; ws.cell(row=r, column=c).border = TB
    r += 1

r += 1; ws.cell(row=r, column=1, value='Key Findings').font = SF; r += 1
findings = [
    f'1. HS6 accuracy: {h6_ok/N*100:.1f}% — Chapter: {ch_ok/N*100:.1f}%, Heading: {h_ok/N*100:.1f}%',
    '2. 72.5% of errors are KEYWORD_MISSING — synonym dictionaries need expansion for AliExpress vocabulary',
    '3. 27.5% are FIELD_DEPENDENT — material field missing (57% availability vs 100% in Amazon test)',
    '4. Jewelry category (100+ items) = 0% accuracy — Ch.71 heading disambiguation needs major expansion',
    '5. Step 2-3 (Chapter) is the primary failure point (61% of errors)',
    '6. This is a REAL independent benchmark — not self-referential like Amazon 50',
    '7. Pipeline designed for seller data (9-field) tested on listing data (~5-field) — expected gap',
    f'8. Code fix potential: {fixes["code_fix_needed"]} items fixable with keyword/rule additions',
]
for f in findings:
    ws.cell(row=r, column=1, value=f).font = NF; r += 1

aw(ws, 8, 60)

# ═══ Sheet: HSCodeComp 632 Detail ═══
ws2 = wb.create_sheet('HSCodeComp Detail')
ws2.sheet_properties.tabColor = '00B050'
dh = ['#','Product Name','Material','Category','Fields','Verified HS6','Pipeline HS6','Ch Match','H Match','HS6 Match','Fail Step','Error Type']
for c, h in enumerate(dh, 1): ws2.cell(row=1, column=c, value=h)
hdr(ws2, 1, len(dh))
for i, res in enumerate(results):
    r = i + 2
    ws2.cell(row=r, column=1, value=res['id']).font = NF
    ws2.cell(row=r, column=2, value=res['product_name'][:45]).font = NF
    ws2.cell(row=r, column=3, value=res['material']).font = NF
    ws2.cell(row=r, column=4, value=res['category_short'][:30]).font = NF
    ws2.cell(row=r, column=5, value=res['available_field_count']).font = NF
    ws2.cell(row=r, column=6, value=res['verified_hs6']).font = NF
    ws2.cell(row=r, column=7, value=res['pipeline_hs6']).font = NF
    for ci, k in enumerate(['chapter_match','heading_match','hs6_match'], 8):
        cell = ws2.cell(row=r, column=ci, value='✅' if res[k] else '❌')
        cell.font = NF; cell.fill = GF if res[k] else RF
    # Find error info
    err = next((e for e in errors_list if e['id'] == res['id']), None)
    ws2.cell(row=r, column=11, value=err['fail_step'] if err else '').font = NF
    ws2.cell(row=r, column=12, value=err['error_type'] if err else '').font = NF
    for c in range(1, len(dh)+1): ws2.cell(row=r, column=c).border = TB
aw(ws2, len(dh), 45)

# ═══ Sheet: HSCodeComp Errors ═══
ws3 = wb.create_sheet('HSCodeComp Errors')
ws3.sheet_properties.tabColor = 'FF4444'
eh = ['#','Product Name','Verified HS6','Pipeline HS6','Fail Step','Error Type','Root Cause','Code Fix?','Fix Description']
for c, h in enumerate(eh, 1): ws3.cell(row=1, column=c, value=h)
hdr(ws3, 1, len(eh))
for i, e in enumerate(errors_list[:300]):  # first 300 errors
    r = i + 2
    ws3.cell(row=r, column=1, value=e['id']).font = NF
    ws3.cell(row=r, column=2, value=e['product_name'][:40]).font = NF
    ws3.cell(row=r, column=3, value=e['verified_hs6']).font = NF
    ws3.cell(row=r, column=4, value=e['pipeline_hs6']).font = NF
    ws3.cell(row=r, column=5, value=e['fail_step']).font = NF
    ws3.cell(row=r, column=6, value=e['error_type']).font = NF
    et_colors = {'KEYWORD_MISSING': OF, 'FIELD_DEPENDENT': PatternFill(start_color='D9D9D9', fill_type='solid'), 'LOGIC_BUG': RF}
    ws3.cell(row=r, column=6).fill = et_colors.get(e['error_type'], PatternFill())
    ws3.cell(row=r, column=7, value=e['root_cause'][:60]).font = NF
    ws3.cell(row=r, column=8, value='YES' if e['code_fix_needed'] else 'No').font = NF
    if e['code_fix_needed']: ws3.cell(row=r, column=8).fill = RF
    ws3.cell(row=r, column=9, value=e['fix_description'][:50]).font = NF
    for c in range(1, len(eh)+1):
        ws3.cell(row=r, column=c).border = TB
        ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
aw(ws3, len(eh), 50)

# ═══ Sheet: HSCodeComp Fixes ═══
ws4 = wb.create_sheet('HSCodeComp Fixes')
ws4.sheet_properties.tabColor = '00B050'
r = 1
ws4.cell(row=r, column=1, value='HSCodeComp — Code Fix Candidates').font = TF; r = 3
ws4.cell(row=r, column=1, value=f'Total errors: {len(errors_list)}').font = BF; r += 1
ws4.cell(row=r, column=1, value=f'KEYWORD_MISSING: {fixes["by_type"].get("KEYWORD_MISSING",0)} — fixable with synonym/keyword additions').font = NF; r += 1
ws4.cell(row=r, column=1, value=f'FIELD_DEPENDENT: {fixes["by_type"].get("FIELD_DEPENDENT",0)} — not fixable (missing input data)').font = NF; r += 2

ws4.cell(row=r, column=1, value='Priority Fix Areas (by error count):').font = SF; r += 1

# Aggregate by chapter for fix priority
ch_errors = {}
for e in errors_list:
    if e['error_type'] == 'KEYWORD_MISSING':
        ch = e['verified_hs6'][:2] if e['verified_hs6'] else '??'
        ch_errors[ch] = ch_errors.get(ch, 0) + 1

fh = ['Chapter', 'Error Count', 'Example Products', 'Fix Type']
for c, h in enumerate(fh, 1): ws4.cell(row=r, column=c, value=h)
hdr(ws4, r, len(fh)); r += 1

for ch, cnt in sorted(ch_errors.items(), key=lambda x: -x[1])[:20]:
    examples = [e['product_name'][:30] for e in errors_list if e['verified_hs6'][:2] == ch and e['error_type'] == 'KEYWORD_MISSING'][:3]
    ws4.cell(row=r, column=1, value=f'Ch.{ch}').font = NF
    ws4.cell(row=r, column=2, value=cnt).font = NF
    ws4.cell(row=r, column=3, value='; '.join(examples)).font = NF
    ws4.cell(row=r, column=4, value='step2-3 chapter mapping + step3 heading keywords').font = NF
    for c in range(1,5): ws4.cell(row=r, column=c).border = TB
    r += 1

aw(ws4, 4, 60)

wb.save(f'{BASE}/POTAL_Ablation_V2.xlsx')
print(f'✅ 4 HSCodeComp sheets added to POTAL_Ablation_V2.xlsx')
