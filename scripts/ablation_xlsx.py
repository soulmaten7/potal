#!/usr/bin/env python3
"""
Amazon 50-product Ablation Test → Excel
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = '/Volumes/soulmaten/POTAL/7field_benchmark'
INPUT_FILE = f'{BASE}/ablation_results.json'
OUTPUT_FILE = f'{BASE}/Amazon_Ablation_Test.xlsx'

with open(INPUT_FILE) as f:
    results = json.load(f)

# Styles
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
LIGHT_GREEN = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='2F5496')
SUBTITLE_FONT = Font(name='Arial', size=12, bold=True)
BOLD_FONT = Font(name='Arial', size=11, bold=True)
NORMAL_FONT = Font(name='Arial', size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def auto_width(ws, max_col, max_width=40):
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def write_result_table(ws, start_row, data_list, include_delta=False):
    """Write a result table with optional delta columns"""
    if include_delta:
        headers = ['Test Name', 'Fields', 'Field Count', 'Section %', 'Chapter %', 'Heading %', 'HS6 %',
                   'ΔSection', 'ΔChapter', 'ΔHeading', 'ΔHS6']
        max_col = 11
    else:
        headers = ['Test Name', 'Fields', 'Field Count', 'Section %', 'Chapter %', 'Heading %', 'HS6 %']
        max_col = 7

    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header(ws, start_row, max_col)

    row = start_row + 1
    baseline = results[0]  # First entry is always baseline

    for r in data_list:
        ws.cell(row=row, column=1, value=r['name']).font = BOLD_FONT if r['name'].startswith('Baseline') else NORMAL_FONT
        fields_str = ', '.join(r.get('fields_used', []))
        ws.cell(row=row, column=2, value=fields_str[:60]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=r['field_count']).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=f"{r['section_pct']}%").font = NORMAL_FONT
        ws.cell(row=row, column=5, value=f"{r['chapter_pct']}%").font = NORMAL_FONT
        ws.cell(row=row, column=6, value=f"{r['heading_pct']}%").font = NORMAL_FONT
        ws.cell(row=row, column=7, value=f"{r['hs6_pct']}%").font = NORMAL_FONT

        if include_delta and r['name'] != 'Baseline (9/9)':
            d_s = r['section_pct'] - baseline['section_pct']
            d_c = r['chapter_pct'] - baseline['chapter_pct']
            d_h = r['heading_pct'] - baseline['heading_pct']
            d_6 = r['hs6_pct'] - baseline['hs6_pct']
            ws.cell(row=row, column=8, value=f"{d_s:+d}%").font = NORMAL_FONT
            ws.cell(row=row, column=9, value=f"{d_c:+d}%").font = NORMAL_FONT
            ws.cell(row=row, column=10, value=f"{d_h:+d}%").font = NORMAL_FONT
            ws.cell(row=row, column=11, value=f"{d_6:+d}%").font = NORMAL_FONT

            # Color delta cells
            for dc in range(8, 12):
                cell = ws.cell(row=row, column=dc)
                cell.border = THIN_BORDER
                val = [d_s, d_c, d_h, d_6][dc-8]
                if val == 0:
                    cell.fill = GREEN_FILL
                elif val >= -10:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = RED_FILL

        # Color accuracy cells
        for ac in range(4, 8):
            cell = ws.cell(row=row, column=ac)
            pct_str = str(cell.value).replace('%', '')
            try:
                pct = int(pct_str)
                if pct >= 95:
                    cell.fill = GREEN_FILL
                elif pct >= 70:
                    cell.fill = YELLOW_FILL
                elif pct >= 40:
                    cell.fill = PatternFill(start_color='FFD9B3', end_color='FFD9B3', fill_type='solid')
                else:
                    cell.fill = RED_FILL
            except:
                pass

        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True)

        row += 1

    return row

wb = Workbook()

# ═══════════════════════════════════════════════
# Sheet 1: Summary
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = 'Summary'

row = 1
ws.cell(row=row, column=1, value='Amazon 50-Product — 9-Field Ablation Test').font = TITLE_FONT
ws.merge_cells('A1:G1')
row += 2

# All results table
ws.cell(row=row, column=1, value='All Test Results (20 combinations)').font = SUBTITLE_FONT
row += 1
row = write_result_table(ws, row, results, include_delta=True)

row += 1
ws.cell(row=row, column=1, value='Key Findings').font = SUBTITLE_FONT
row += 1

findings = [
    '1. material is the MOST CRITICAL field — removing it drops Section accuracy from 100% to 48% (-52%)',
    '2. category is the SECOND most important — removing it drops Heading from 100% to 60% (-40%)',
    '3. product_name + material + category = 3 fields achieve 100% Section/Chapter/Heading, 98% HS6',
    '4. origin_country has ZERO impact on classification accuracy (expected — it affects duty rates, not HS codes)',
    '5. processing, composition, weight_spec, price have ZERO impact at the heading level for these 50 items',
    '6. description adds only +2% at HS6 level (100% vs 98%)',
    '7. product_name alone achieves only 36% Section, 22% Chapter — insufficient without material',
    '8. Removing both material + category = worst case: 36% Section, 22% Chapter, 22% Heading',
    '9. The "magic 3" fields (name + material + category) are the minimum for reliable classification',
]
for finding in findings:
    ws.cell(row=row, column=1, value=finding).font = NORMAL_FONT
    row += 1

row += 1
ws.cell(row=row, column=1, value='Field Impact Ranking (most → least important)').font = SUBTITLE_FONT
row += 1

rank_headers = ['Rank', 'Field', 'Impact Grade', 'HS6 Drop When Removed', 'Role']
for c, h in enumerate(rank_headers, 1):
    ws.cell(row=row, column=c, value=h)
style_header(ws, row, len(rank_headers))
row += 1

rankings = [
    ('1', 'material', 'CRITICAL', '-60%', 'Section selection (S1-S21), Chapter refinement'),
    ('2', 'category', 'CRITICAL', '-40%', 'Heading selection, Section override (clothing vs jewelry)'),
    ('3', 'product_name', 'HIGH', '-16%', 'Heading keywords, subheading matching'),
    ('4', 'description', 'LOW', '-2%', 'Subheading disambiguation (only at HS6 level)'),
    ('5', 'origin_country', 'NONE', '0%', 'No impact on HS classification (affects duty rates only)'),
    ('6', 'processing', 'NONE', '0%', 'No impact for consumer goods (affects industrial/food)'),
    ('7', 'composition', 'NONE', '0%', 'No impact for consumer goods (affects alloys/textiles)'),
    ('8', 'weight_spec', 'NONE', '0%', 'No impact for consumer goods (affects metals/chemicals)'),
    ('9', 'price', 'NONE', '0%', 'No impact on HS6 (affects HS10 price breaks only)'),
]

grade_colors = {
    'CRITICAL': PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid'),
    'HIGH': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
    'LOW': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'NONE': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
}

for rank_data in rankings:
    for c, val in enumerate(rank_data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(name='Arial', size=11, bold=(c <= 2), color='FFFFFF' if rank_data[2] == 'CRITICAL' else '000000')
        cell.border = THIN_BORDER
    ws.cell(row=row, column=3).fill = grade_colors.get(rank_data[2], GREEN_FILL)
    row += 1

auto_width(ws, 11, 50)

# ═══════════════════════════════════════════════
# Sheet 2: Round 1
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet('Round 1 — Remove 1')

row = 1
ws2.cell(row=row, column=1, value='Round 1: Remove 1 Field at a Time').font = TITLE_FONT
ws2.merge_cells('A1:G1')
row += 2

# Round 1 data: baseline + 9 remove-1 tests (indices 0..9)
round1_data = results[0:10]
# Sort by HS6 drop (ascending = biggest drop first)
baseline_hs6 = results[0]['hs6_pct']
sorted_r1 = [results[0]] + sorted(results[1:10], key=lambda r: r['hs6_pct'])
row = write_result_table(ws2, row, sorted_r1, include_delta=True)

row += 1
ws2.cell(row=row, column=1, value='Sorted by HS6 accuracy drop (worst first)').font = NORMAL_FONT

auto_width(ws2, 11, 50)

# ═══════════════════════════════════════════════
# Sheet 3: Round 2
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet('Round 2 — Minimum')

row = 1
ws3.cell(row=row, column=1, value='Round 2: Minimum Field Combinations').font = TITLE_FONT
ws3.merge_cells('A1:G1')
row += 2

# Round 2 data: indices 10..15
round2_data = [results[0]] + results[10:16]
row = write_result_table(ws3, row, round2_data, include_delta=True)

row += 1
ws3.cell(row=row, column=1, value='Key Insight').font = SUBTITLE_FONT
row += 1
insights = [
    'product_name alone: 36% Section — practically useless',
    'product_name + material: 96% Section, 58% HS6 — material is transformative',
    'product_name + category: 48% Section, 24% HS6 — category alone insufficient',
    'product_name + material + category: 100% Section/Chapter/Heading, 98% HS6 — the MAGIC 3',
    'Adding origin_country to magic 3: no change (0% improvement)',
    'Adding description to magic 3: +2% HS6 (98→100%) — marginal gain',
    '',
    'MINIMUM FOR RELIABLE CLASSIFICATION: product_name + material + category (3 fields)',
    'RECOMMENDED: product_name + material + category + description (4 fields)',
]
for ins in insights:
    ws3.cell(row=row, column=1, value=ins).font = BOLD_FONT if 'MINIMUM' in ins or 'RECOMMENDED' in ins else NORMAL_FONT
    row += 1

auto_width(ws3, 11, 50)

# ═══════════════════════════════════════════════
# Sheet 4: Round 3
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet('Round 3 — Remove 2')

row = 1
ws4.cell(row=row, column=1, value='Round 3: Remove 2 Fields Simultaneously').font = TITLE_FONT
ws4.merge_cells('A1:G1')
row += 2

round3_data = [results[0]] + results[16:21]
row = write_result_table(ws4, row, round3_data, include_delta=True)

row += 1
ws4.cell(row=row, column=1, value='Key Insight').font = SUBTITLE_FONT
row += 1
ws4.cell(row=row, column=1, value='material + category together = catastrophic (-78% HS6). Each complements the other.').font = NORMAL_FONT
row += 1
ws4.cell(row=row, column=1, value='processing + composition removal = 0% impact. weight_spec + price removal = 0% impact.').font = NORMAL_FONT
row += 1
ws4.cell(row=row, column=1, value='Only material and category matter at the 4-digit heading level for consumer goods.').font = NORMAL_FONT

auto_width(ws4, 11, 50)

# ═══════════════════════════════════════════════
# Sheet 5: Field Impact Analysis
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet('Field Impact Analysis')

row = 1
ws5.cell(row=row, column=1, value='Per-Field Impact Analysis').font = TITLE_FONT
ws5.merge_cells('A1:H1')
row += 2

fi_headers = ['Field', 'Required?', 'Impact Grade', 'Section Δ', 'Chapter Δ', 'Heading Δ', 'HS6 Δ',
              'Affected Pipeline Step', 'Seller Message']
for c, h in enumerate(fi_headers, 1):
    ws5.cell(row=row, column=c, value=h)
style_header(ws5, row, len(fi_headers))
row += 1

field_analysis = [
    ('material', 'Required', 'CRITICAL', '-52%', '-60%', '-60%', '-60%',
     'Step 2-1 (Section), Step 2-3 (Chapter), Step 3 (Heading tiebreaker)',
     '"Adding material increases accuracy by +60%. This is the most important field for HS classification."'),
    ('category', 'Recommended', 'CRITICAL', '-4%', '-22%', '-40%', '-40%',
     'Step 2-1 (Section override), Step 3 (Heading synonym dict)',
     '"Adding your product category increases accuracy by +40%. E.g., \'Clothing > Shirts > T-Shirts\'"'),
    ('product_name', 'Required', 'HIGH', '-10%', '-16%', '-16%', '-16%',
     'Step 3 (Heading keywords), Step 4 (Subheading synonyms)',
     '"Product name provides essential keywords for precise classification."'),
    ('description', 'Optional', 'LOW', '0%', '0%', '0%', '-2%',
     'Step 4 (Subheading only)',
     '"Adding description can improve 6-digit precision by up to +2%."'),
    ('origin_country', 'Required*', 'NONE', '0%', '0%', '0%', '0%',
     'Step 5-7 (Country routing, NOT classification)',
     '"Required for duty rate calculation, but does not affect HS code classification."'),
    ('processing', 'Optional', 'NONE', '0%', '0%', '0%', '0%',
     'Step 2-2 (Chapter hint for food/textiles)',
     '"Useful for industrial/food products. No impact for typical consumer goods."'),
    ('composition', 'Optional', 'NONE', '0%', '0%', '0%', '0%',
     'Step 4 (Alloy detection, textile blend)',
     '"Useful for alloys, textile blends. No impact for single-material consumer goods."'),
    ('weight_spec', 'Optional', 'NONE', '0%', '0%', '0%', '0%',
     'Step 4 (Subheading weight thresholds)',
     '"Useful for metals/chemicals with weight-based tariff breaks."'),
    ('price', 'Optional', 'NONE', '0%', '0%', '0%', '0%',
     'Step 6 (Price break at HS10 level)',
     '"Used for 10-digit price breaks (e.g., \'valued over $5\'), not 6-digit classification."'),
]

for fa in field_analysis:
    for c, val in enumerate(fa, 1):
        cell = ws5.cell(row=row, column=c, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
    # Color the grade
    grade = fa[2]
    ws5.cell(row=row, column=3).fill = grade_colors.get(grade, GREEN_FILL)
    if grade == 'CRITICAL':
        ws5.cell(row=row, column=3).font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    row += 1

row += 2
ws5.cell(row=row, column=1, value='Seller-Facing Tier System').font = SUBTITLE_FONT
row += 1

tier_headers = ['Tier', 'Fields Provided', 'Expected Accuracy', 'Seller Experience']
for c, h in enumerate(tier_headers, 1):
    ws5.cell(row=row, column=c, value=h)
style_header(ws5, row, len(tier_headers))
row += 1

tiers = [
    ('Tier 1: Basic', 'product_name + material + origin', '58% HS6', 'Minimum viable — many incorrect classifications'),
    ('Tier 2: Good', 'Tier 1 + category', '98% HS6', 'Highly accurate — recommended for most sellers'),
    ('Tier 3: Best', 'Tier 2 + description', '100% HS6', 'Perfect accuracy — gold standard'),
    ('Tier 4: Full', 'All 9 fields', '100% HS6', 'Maximum data — no additional accuracy gain but enables HS10 price breaks'),
]
for tier_data in tiers:
    for c, val in enumerate(tier_data, 1):
        cell = ws5.cell(row=row, column=c, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
    row += 1

auto_width(ws5, 9, 60)

# Save
wb.save(OUTPUT_FILE)
print(f'✅ Excel saved: {OUTPUT_FILE}')
print(f'   5 sheets: Summary, Round 1, Round 2, Round 3, Field Impact Analysis')
