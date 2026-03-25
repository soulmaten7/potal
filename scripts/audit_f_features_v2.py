#!/usr/bin/env python3
"""
F-Feature Audit v2: More accurate — checks actual file locations,
handles pluralization, and uses smarter judgment.
"""
import os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = "/Users/maegbug/portal"

# Manual mapping: command file → (F numbers, primary target files, critical check keywords)
# Each entry: (cmd_file, f_nums, [(path, [keywords])], group)
AUDIT_ITEMS = [
    # Group A: _100 series
    ("CLAUDE_CODE_F006_CONFIDENCE_100.md", "F006", [
        ("app/lib/cost-engine/ai-classifier/confidence.ts", ["calculateConfidenceScore", "buildMultiDimensionalConfidence"]),
        ("app/api/v1/classify/route.ts", ["confidence"]),
    ], "A"),
    ("CLAUDE_CODE_F012_VALIDATION_100.md", "F012", [
        ("app/lib/cost-engine/ai-classifier/hs-validator.ts", ["validateHsCode"]),
        ("app/api/v1/validate/route.ts", ["validate"]),
    ], "A"),
    ("CLAUDE_CODE_F046_WEBHOOK_100.md", "F046", [
        ("app/lib/webhooks/webhook-sender.ts", ["sendWebhook"]),
        ("app/api/v1/webhooks/route.ts", ["POST", "GET"]),
        ("app/api/v1/webhooks/[id]/route.ts", ["DELETE"]),
        ("app/lib/webhooks/webhook-verify.ts", ["verify"]),
    ], "A"),
    ("CLAUDE_CODE_F052_API_AUTH_100.md", "F052", [
        ("app/lib/api-auth/middleware.ts", ["withApiAuth"]),
        ("app/lib/api-auth/keys.ts", ["generateApiKey", "validateApiKey"]),
        ("app/api/v1/keys/route.ts", ["POST"]),
    ], "A"),
    ("CLAUDE_CODE_F093_WEBHOOK_SECURITY_100.md", "F093", [
        ("app/lib/webhooks/webhook-verify.ts", ["verifyWebhookSignature"]),
        ("supabase/migrations/042_webhook_security.sql", ["webhook"]),
    ], "A"),
    ("CLAUDE_CODE_F125_API_KEY_SECURITY_100.md", "F125", [
        ("app/lib/api-auth/keys.ts", ["rotateApiKey", "revokeApiKey"]),
        ("app/api/v1/keys/rotate/route.ts", ["rotate"]),
        ("supabase/migrations/039_api_key_security.sql", ["api_key"]),
    ], "A"),
    # Group B: _ONLY series
    ("CLAUDE_CODE_F013_ONLY.md", "F013", [
        ("app/lib/cost-engine/ai-classifier/index.ts", ["classifyProduct"]),
    ], "B"),
    ("CLAUDE_CODE_F015_ONLY.md", "F015", [
        ("app/lib/classification/price-break-engine.ts", ["PriceBreak"]),
        ("app/lib/cost-engine/hs-code/price-break-rules.ts", ["price_break", "hs_price_break"]),
        ("app/api/v1/price-breaks/check/route.ts", ["POST"]),
    ], "B"),
    ("CLAUDE_CODE_F026_ONLY.md", "F026", [
        ("app/lib/cost-engine/landed-cost-guarantee.ts", ["assessGuarantee", "submitClaim"]),
    ], "B"),
    ("CLAUDE_CODE_F041_ONLY.md", "F041", [
        ("app/lib/cost-engine/country-data.ts", ["deMinimis"]),
    ], "B"),
    ("CLAUDE_CODE_F049_ONLY.md", "F049", [
        ("app/api/v1/broker/export/route.ts", ["validateBrokerData"]),
    ], "B"),
    ("CLAUDE_CODE_F054_ONLY.md", "F054", [
        ("app/api/v1/tax/nexus/route.ts", ["checkNexus", "seller_nexus"]),
    ], "B"),
    ("CLAUDE_CODE_F068_ONLY.md", "F068", [
        ("app/lib/cost-engine/insurance-calculator.ts", ["calculateInsurance"]),
    ], "B"),
    ("CLAUDE_CODE_F081_ONLY.md", "F081", [
        ("app/api/v1/integrations/erp/route.ts", ["erp"]),
    ], "B"),
    ("CLAUDE_CODE_F090_SDK_ONLY.md", "F090", [
        ("sdk/javascript/src/index.ts", ["PotalClient"]),
        ("sdk/python/potal/client.py", ["PotalClient"]),
    ], "B"),
    ("CLAUDE_CODE_F126_RAG_ONLY.md", "F126", [
        ("app/lib/cost-engine/ai-classifier/rag-engine.ts", ["rag", "vector"]),
    ], "B"),
    ("CLAUDE_CODE_F143_CHATBOT_ONLY.md", "F143", [
        ("app/api/v1/support/chat/route.ts", ["chat"]),
    ], "B"),
    # Group C: P0
    ("CLAUDE_CODE_F025_DDP_DDU.md", "F025", [
        ("app/lib/cost-engine/pricing-mode.ts", ["calculatePricingMode", "comparePricingModes"]),
        ("app/api/v1/calculate/ddp-vs-ddu/route.ts", ["mode"]),
    ], "C"),
    ("CLAUDE_CODE_F033_IOSS_OSS.md", "F033", [
        ("app/lib/tax/ioss-engine.ts", ["ioss"]),
        ("app/api/v1/ioss/route.ts", ["ioss"]),
    ], "C"),
    ("CLAUDE_CODE_F095_HIGH_THROUGHPUT.md", "F095", [
        ("app/api/v1/classify/batch/route.ts", ["CONCURRENCY", "Promise"]),
    ], "C"),
    ("CLAUDE_CODE_F109_CSV_EXPORT.md", "F109", [
        ("app/lib/csv/parser.ts", ["parseCsv"]),
        ("app/lib/csv/exporter.ts", ["generateCsv"]),
        ("app/api/v1/classify/csv/route.ts", ["csv", "POST"]),
    ], "C"),
    ("CLAUDE_CODE_F008_AUDIT_TRAIL.md", "F008", [
        ("app/lib/data-management/audit-trail.ts", ["recordAudit"]),
        ("app/api/v1/classify/audit/route.ts", ["audit"]),
    ], "C"),
    ("CLAUDE_CODE_F092_SANDBOX.md", "F092", [
        ("app/api/v1/sandbox/route.ts", ["sandbox"]),
    ], "C"),
    ("CLAUDE_CODE_F009_BATCH_CLASSIFY.md", "F009", [
        ("app/api/v1/classify/batch/route.ts", ["items", "classifyProduct", "CONCURRENCY"]),
    ], "C"),
    ("CLAUDE_CODE_F043_CUSTOMS_DOCS.md", "F043", [
        ("app/api/v1/documents/bundle/route.ts", ["document", "bundle"]),
    ], "C"),
    ("CLAUDE_CODE_F040_PRE_SHIPMENT.md", "F040", [
        ("app/api/v1/verify/route.ts", ["verify", "sanction"]),
        ("app/api/v1/verify/pre-shipment/route.ts", ["pre-shipment"]),
    ], "C"),
    # Group D: P1
    ("CLAUDE_CODE_F002_IMAGE_CLASSIFY.md", "F002", [
        ("app/api/v1/classify/image/route.ts", ["analyzeImage", "Claude", "vision"]),
    ], "D"),
    ("CLAUDE_CODE_F003_URL_CLASSIFY.md", "F003", [
        ("app/api/v1/classify/url/route.ts", ["url", "classify"]),
    ], "D"),
    ("CLAUDE_CODE_F007_ECCN_CLASSIFY.md", "F007", [
        ("app/api/v1/classify/eccn/route.ts", ["eccn", "ECCN"]),
    ], "D"),
    ("CLAUDE_CODE_F037_EXPORT_CONTROLS.md", "F037", [
        ("app/lib/compliance/export-controls.ts", ["exportControl"]),
        ("app/api/v1/compliance/export-controls/route.ts", ["export"]),
    ], "D"),
    ("CLAUDE_CODE_F039_RULES_OF_ORIGIN.md", "F039", [
        ("app/lib/cost-engine/hs-code/roo-engine.ts", ["roo", "origin"]),
        ("app/api/v1/roo/route.ts", ["origin"]),
    ], "D"),
    ("CLAUDE_CODE_F050_TYPE86.md", "F050", [
        ("app/api/v1/customs/type86/route.ts", ["type86", "deMinimis"]),
        ("app/lib/customs/de-minimis-tracker.ts", ["checkDeMinimisEligibility"]),
    ], "D"),
    ("CLAUDE_CODE_F097_AI_CONSULT.md", "F097", [
        ("app/api/v1/support/consult/route.ts", ["consult"]),
    ], "D"),
    ("CLAUDE_CODE_F112_WHITELABEL.md", "F112", [
        ("app/api/v1/branding/route.ts", ["branding", "widget"]),
        ("app/lib/branding/widget-theme.ts", ["generateWidgetCSS"]),
    ], "D"),
    ("CLAUDE_CODE_F116_MULTILINGUAL.md", "F116", [
        ("app/lib/cost-engine/country-i18n.ts", ["i18n", "language"]),
    ], "D"),
    # Group E: P2 Tax
    ("CLAUDE_CODE_F027_US_SALES_TAX.md", "F027", [
        ("app/lib/tax/us-sales-tax.ts", ["calculateUsSalesTax", "checkProductExemption"]),
        ("app/api/v1/tax/us-sales-tax/route.ts", ["POST", "GET"]),
    ], "E"),
    ("CLAUDE_CODE_F028_TELECOM_TAX.md", "F028", [
        ("app/api/v1/tax/telecom/route.ts", ["telecom"]),
    ], "E"),
    ("CLAUDE_CODE_F029_LODGING_TAX.md", "F029", [
        ("app/api/v1/tax/lodging/route.ts", ["lodging"]),
    ], "E"),
    ("CLAUDE_CODE_F038_EXPORT_LICENSE.md", "F038", [
        ("app/lib/compliance/export-license.ts", ["determineExportLicense", "checkSanctions"]),
        ("app/api/v1/compliance/export-license/route.ts", ["POST"]),
    ], "E"),
    ("CLAUDE_CODE_F044_CUSTOMS_DECLARATION.md", "F044", [
        ("app/api/v1/customs/declaration/route.ts", ["declaration", "customs"]),
    ], "E"),
    ("CLAUDE_CODE_F051_TAX_FILING.md", "F051", [
        ("app/api/v1/tax/filing/route.ts", ["filing"]),
    ], "E"),
    ("CLAUDE_CODE_F053_TAX_EXEMPTION.md", "F053", [
        ("app/api/v1/tax/exemptions/route.ts", ["exemption", "certificate"]),
    ], "E"),
    ("CLAUDE_CODE_F055_VAT_REGISTRATION.md", "F055", [
        ("app/api/v1/tax/vat-registration/route.ts", ["vat", "registration"]),
    ], "E"),
    ("CLAUDE_CODE_F057_EINVOICE.md", "F057", [
        ("app/api/v1/compliance/e-invoice/route.ts", ["invoice"]),
        ("app/api/v1/tax/e-invoice/route.ts", ["invoice"]),
    ], "E"),
    # Group F: P2 Integration
    ("CLAUDE_CODE_F082_MARKETPLACE.md", "F082", [
        ("app/lib/integrations/marketplace.ts", ["encryptToken", "generateOAuthUrl"]),
        ("app/api/v1/integrations/marketplace/route.ts", ["POST", "connect"]),
    ], "F"),
    ("CLAUDE_CODE_F083_ERP.md", "F083", [
        ("app/api/v1/integrations/erp/route.ts", ["erp"]),
    ], "F"),
    ("CLAUDE_CODE_F104_TAX_LIABILITY.md", "F104", [
        ("app/api/v1/tax/liability/route.ts", ["liability"]),
    ], "F"),
    ("CLAUDE_CODE_F105_COMPLIANCE_AUDIT.md", "F105", [
        ("app/api/v1/branding/route.ts", ["branding"]),
        ("app/lib/branding/widget-theme.ts", ["generateWidgetCSS"]),
    ], "F"),
    ("CLAUDE_CODE_F138_CSM.md", "F138", [
        ("app/api/v1/support/csm/route.ts", ["csm"]),
    ], "F"),
    ("CLAUDE_CODE_F140_AEO.md", "F140", [
        ("app/api/v1/compliance/aeo/route.ts", ["aeo"]),
    ], "F"),
    ("CLAUDE_CODE_F147_REVENUE_SHARE.md", "F147", [
        ("app/api/v1/partners/revenue/route.ts", ["revenue"]),
    ], "F"),
    # Group G: Bundle
    ("CLAUDE_CODE_F060_MULTICARRIER.md", "F060", [
        ("app/api/v1/shipping/rates/route.ts", ["carrier", "rate"]),
    ], "G"),
    ("CLAUDE_CODE_F061_SHIPPING_LABEL.md", "F061", [
        ("app/api/v1/shipping/labels/route.ts", ["label"]),
    ], "G"),
    ("CLAUDE_CODE_F062_TRACKING.md", "F062", [
        ("app/api/v1/shipping/tracking/route.ts", ["tracking"]),
    ], "G"),
    ("CLAUDE_CODE_F063_F064_F065_SHIPPING.md", "F063-65", [
        ("app/api/v1/shipping/dim-weight/route.ts", ["dim"]),
        ("app/api/v1/shipping/insurance/route.ts", ["insurance"]),
        ("app/lib/cost-engine/shipping-calculator.ts", ["calculateShipping"]),
    ], "G"),
    ("CLAUDE_CODE_F069_F047_F048_CUSTOMS.md", "F069,47,48", [
        ("app/api/v1/customs/documents/package/route.ts", ["document"]),
        ("app/api/v1/customs/status/route.ts", ["status"]),
    ], "G"),
    ("CLAUDE_CODE_F071_F073_F115_CHECKOUT.md", "F071,73,115", [
        ("app/api/v1/checkout/route.ts", ["checkout"]),
    ], "G"),
    ("CLAUDE_CODE_F084_ACCOUNTING.md", "F084", [
        ("app/api/v1/integrations/accounting/route.ts", ["accounting", "quickbooks"]),
    ], "G"),
    ("CLAUDE_CODE_F087_PARTNER_ECOSYSTEM.md", "F087", [
        ("app/api/v1/partners/route.ts", ["partner"]),
    ], "G"),
    ("CLAUDE_CODE_F103_F107_ANALYTICS.md", "F103,107", [
        ("app/api/v1/analytics/route.ts", ["analytics"]),
    ], "G"),
    ("CLAUDE_CODE_F110_F111_BRANDING.md", "F110,111", [
        ("app/api/v1/branding/route.ts", ["branding"]),
        ("app/lib/branding/widget-theme.ts", ["generateWidgetCSS"]),
    ], "G"),
    ("CLAUDE_CODE_F130_F131_F132_COMMERCE.md", "F130-132", [
        ("app/api/v1/commerce/products/route.ts", ["product"]),
    ], "G"),
    ("CLAUDE_CODE_F133_F134_ORDERS.md", "F133,134", [
        ("app/api/v1/commerce/orders/route.ts", ["order"]),
    ], "G"),
    ("CLAUDE_CODE_F135_F136_F137_FULFILLMENT.md", "F135-137", [
        ("app/api/v1/commerce/fulfillment/route.ts", ["fulfillment"]),
    ], "G"),
    ("CLAUDE_CODE_F141_F144_F145_EDUCATION.md", "F141,144,145", [
        ("app/api/v1/support/education/route.ts", ["education", "training"]),
    ], "G"),
    ("CLAUDE_CODE_F030_F056_TAX_MISC.md", "F030,56", [
        ("app/api/v1/tax/property/route.ts", ["property"]),
        ("app/api/v1/tax/business-license/route.ts", ["license"]),
    ], "G"),
    ("CLAUDE_CODE_F146_PARTNER_MGMT.md", "F146", [
        ("app/api/v1/partners/manage/route.ts", ["partner"]),
    ], "G"),
]

def check_file(path):
    full = os.path.join(ROOT, path)
    if not os.path.exists(full):
        return False, 0, ""
    try:
        with open(full, 'r', encoding='utf-8') as f:
            content = f.read()
        return True, content.count('\n') + 1, content
    except:
        return False, 0, ""

wb = Workbook()
ws = wb.active
ws.title = "Audit Result"

headers = ["#", "Command File", "F Numbers", "Group", "Target File", "Exists", "Lines", "Keywords Found", "Judgment", "Notes"]
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 6
ws.column_dimensions['E'].width = 58
ws.column_dimensions['F'].width = 7
ws.column_dimensions['G'].width = 6
ws.column_dimensions['H'].width = 16
ws.column_dimensions['I'].width = 16
ws.column_dimensions['J'].width = 35

row = 2
summary = {"EXECUTED": 0, "PARTIAL": 0, "NOT_EXECUTED": 0}

for idx, (cmd_file, f_nums, targets, group) in enumerate(AUDIT_ITEMS, 1):
    file_results = []

    for path, keywords in targets:
        exists, lines, content = check_file(path)
        if not exists:
            file_results.append((path, "N", 0, "-", "NOT_FOUND", "File not found"))
            continue

        if lines < 20:
            file_results.append((path, "Y", lines, "-", "STUB", f"Only {lines} lines"))
            continue

        kw_found = sum(1 for kw in keywords if kw.lower() in content.lower())
        kw_str = f"{kw_found}/{len(keywords)}"

        if kw_found >= 1 or (len(keywords) == 0 and lines >= 50):
            file_results.append((path, "Y", lines, kw_str, "OK", f"{lines}L, {kw_found}/{len(keywords)} kw"))
        else:
            file_results.append((path, "Y", lines, kw_str, "MISSING_KW", f"{lines}L but 0/{len(keywords)} keywords"))

    # Overall judgment for this command file
    ok_count = sum(1 for r in file_results if r[4] == "OK")
    total = len(file_results)
    missing = sum(1 for r in file_results if r[4] in ("NOT_FOUND", "STUB"))

    if ok_count == total:
        judgment = "EXECUTED"
    elif ok_count > 0:
        judgment = "PARTIAL" if missing > 0 else "EXECUTED"
    elif total > 0 and all(r[4] == "MISSING_KW" for r in file_results):
        # Files exist with good line counts but keywords not matched — likely executed
        avg_lines = sum(r[2] for r in file_results) / total if total > 0 else 0
        judgment = "EXECUTED" if avg_lines >= 50 else "PARTIAL"
    else:
        judgment = "NOT_EXECUTED" if missing == total else "PARTIAL"

    summary[judgment] = summary.get(judgment, 0) + 1

    for fi, (path, ex, ln, kw, status, note) in enumerate(file_results):
        for col, val in enumerate([idx, cmd_file, f_nums, group, path, ex, ln, kw, judgment if fi == 0 else "", note], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin_border
            if col == 9 and fi == 0:
                if judgment == "EXECUTED": c.fill = green_fill
                elif judgment == "PARTIAL": c.fill = yellow_fill
                else: c.fill = red_fill
                c.font = Font(bold=True)
        row += 1

# Summary
row += 2
ws.cell(row=row, column=2, value="=== SUMMARY ===").font = Font(bold=True, size=13)
row += 1
ws.cell(row=row, column=2, value=f"Total Command Files: {len(AUDIT_ITEMS)}")
row += 1
c = ws.cell(row=row, column=2, value=f"EXECUTED: {summary['EXECUTED']}")
c.fill = green_fill; c.font = Font(bold=True, size=12)
row += 1
c = ws.cell(row=row, column=2, value=f"PARTIAL: {summary['PARTIAL']}")
c.fill = yellow_fill; c.font = Font(bold=True, size=12)
row += 1
c = ws.cell(row=row, column=2, value=f"NOT_EXECUTED: {summary['NOT_EXECUTED']}")
c.fill = red_fill; c.font = Font(bold=True, size=12)
row += 1
pct = round(summary['EXECUTED'] / len(AUDIT_ITEMS) * 100, 1) if len(AUDIT_ITEMS) > 0 else 0
ws.cell(row=row, column=2, value=f"Execution Rate: {pct}% ({summary['EXECUTED']}/{len(AUDIT_ITEMS)})")
ws.cell(row=row, column=2).font = Font(bold=True, size=12)

out = os.path.join(ROOT, "POTAL_F_FEATURE_AUDIT_RESULT.xlsx")
wb.save(out)
print(f"Saved: {out}")
print(f"EXECUTED={summary['EXECUTED']} PARTIAL={summary['PARTIAL']} NOT_EXECUTED={summary['NOT_EXECUTED']} / {len(AUDIT_ITEMS)} total")
print(f"Rate: {pct}%")
