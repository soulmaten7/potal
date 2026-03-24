#!/usr/bin/env python3
"""Generate POTAL_Platform_Product_Fields_Raw.xlsx — NO POTAL mapping"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = '/Volumes/soulmaten/POTAL/platform_field_mapping'
p1 = json.load(open(f'{BASE}/phase1_global_platforms.json'))
p2 = json.load(open(f'{BASE}/phase2_regional_platforms.json'))
p3 = json.load(open(f'{BASE}/phase3_b2b_platforms.json'))
p4 = json.load(open(f'{BASE}/phase4_export_formats.json'))
OUTPUT = f'{BASE}/POTAL_Platform_Product_Fields_Raw.xlsx'

HF = PatternFill(start_color='2F5496', fill_type='solid')
HFo = Font(name='Arial', size=11, bold=True, color='FFFFFF')
TF = Font(name='Arial', size=14, bold=True, color='2F5496')
SF = Font(name='Arial', size=12, bold=True)
BF = Font(name='Arial', size=11, bold=True)
NF = Font(name='Arial', size=11)
TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
GF = PatternFill(start_color='C6EFCE', fill_type='solid')
RF = PatternFill(start_color='FFC7CE', fill_type='solid')
YF = PatternFill(start_color='FFEB9C', fill_type='solid')

def hdr(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HF; cell.font = HFo; cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = TB

def aw(ws, ncol, mx=45):
    for c in range(1, ncol+1):
        ml = max((len(str(cell.value or '')) for cell in ws[get_column_letter(c)]), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(ml+2, mx)

def write_platform_fields(ws, title, platforms_data, start_row=1):
    r = start_row
    ws.cell(row=r, column=1, value=title).font = TF
    r += 2
    headers = ['Platform', 'HQ', 'Main Sellers', 'Field Name (API)', 'Field Name (UI)', 'Type', 'Description', 'Mandatory', 'Example']
    nc = len(headers)
    for c, h in enumerate(headers, 1): ws.cell(row=r, column=c, value=h)
    hdr(ws, r, nc); r += 1

    for pname, pdata in platforms_data.items():
        if pname == 'metadata' or pname == 'b2b_vs_b2c_comparison': continue
        hq = pdata.get('hq_country', '')
        sellers = pdata.get('main_sellers', '')[:40]
        fields = pdata.get('product_fields', [])
        for f in fields:
            ws.cell(row=r, column=1, value=pname.replace('_', ' ')).font = NF
            ws.cell(row=r, column=2, value=hq).font = NF
            ws.cell(row=r, column=3, value=sellers).font = NF
            ws.cell(row=r, column=4, value=f.get('api_name', f.get('api', f.get('field', '')))[:40]).font = NF
            ws.cell(row=r, column=5, value=f.get('ui_name', f.get('name_en', ''))[:30]).font = NF
            ws.cell(row=r, column=6, value=f.get('type', '')[:20]).font = NF
            ws.cell(row=r, column=7, value=f.get('description', '')[:60]).font = NF
            ws.cell(row=r, column=8, value=str(f.get('mandatory', ''))[:10]).font = NF
            ws.cell(row=r, column=9, value=f.get('example', '')[:30]).font = NF
            for c in range(1, nc+1): ws.cell(row=r, column=c).border = TB; ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
            r += 1
    aw(ws, nc)
    return r

wb = Workbook()

# Sheet 1: Global Platforms
ws1 = wb.active; ws1.title = 'Global Platforms (10)'
write_platform_fields(ws1, 'Global E-commerce Platforms — Product Data Fields', p1)

# Sheet 2: Regional Platforms
ws2 = wb.create_sheet('Regional Platforms (15)')
write_platform_fields(ws2, 'Regional E-commerce Platforms — Product Data Fields', p2)

# Sheet 3: B2B Platforms
ws3 = wb.create_sheet('B2B Platforms (5)')
write_platform_fields(ws3, 'B2B Platforms — Product Data Fields', p3)

# Sheet 4: Export Formats
ws4 = wb.create_sheet('Export Formats')
r = 1; ws4.cell(row=r, column=1, value='Product Data Export File Formats').font = TF; r = 3
h4 = ['Platform', 'Format', 'Column Name', 'Material Column?', 'Notes']
for c, h in enumerate(h4, 1): ws4.cell(row=r, column=c, value=h)
hdr(ws4, r, len(h4)); r += 1
for pname, pdata in p4.items():
    if pname == 'metadata': continue
    fmt = pdata.get('format', '')
    cols = pdata.get('columns', pdata.get('common_columns', []))
    mat = pdata.get('material_column', 'N/A')
    notes = '; '.join(pdata.get('notes', []))[:60] if isinstance(pdata.get('notes'), list) else str(pdata.get('notes', ''))[:60]
    for col in cols[:20]:
        ws4.cell(row=r, column=1, value=pname.replace('_', ' ')).font = NF
        ws4.cell(row=r, column=2, value=fmt[:30]).font = NF
        ws4.cell(row=r, column=3, value=col[:40]).font = NF
        is_mat = 'material' in col.lower() or '材质' in col or '材料' in col
        ws4.cell(row=r, column=4, value='YES' if is_mat else '').font = NF
        if is_mat: ws4.cell(row=r, column=4).fill = GF
        ws4.cell(row=r, column=5, value='').font = NF
        for c in range(1, len(h4)+1): ws4.cell(row=r, column=c).border = TB
        r += 1
aw(ws4, len(h4))

# Sheet 5: All 30 Platforms Summary
ws5 = wb.create_sheet('30 Platforms Summary')
r = 1; ws5.cell(row=r, column=1, value='All 30 Platforms — Product Field Summary').font = TF; r = 3
h5 = ['Platform', 'HQ', 'Type', 'Total Fields', 'Has Material?', 'Material Location', 'Has Composition?', 'Has Processing?', 'Has Category?', 'Has Weight?', 'Has Origin?']
for c, h in enumerate(h5, 1): ws5.cell(row=r, column=c, value=h)
hdr(ws5, r, len(h5)); r += 1

all_platforms = []
for src, ptype in [(p1, 'Global B2C'), (p2, 'Regional B2C'), (p3, 'B2B')]:
    for pname, pdata in src.items():
        if pname in ('metadata', 'b2b_vs_b2c_comparison'): continue
        fields = pdata.get('product_fields', [])
        mat_loc = pdata.get('material_location', 'N/A')
        has_mat = 'DEDICATED' in mat_loc.upper() or 'dedicated' in mat_loc.lower() or 'materials[]' in mat_loc
        has_comp = 'composition' in str(pdata.get('composition_location', '')).lower() and 'no' not in str(pdata.get('composition_location', '')).lower()[:5]
        has_proc = 'processing' in str(pdata.get('processing_location', '')).lower() and 'no' not in str(pdata.get('processing_location', '')).lower()[:5]
        has_cat = any('categ' in str(f.get('api_name', '')).lower() or 'categ' in str(f.get('description', '')).lower() for f in fields)
        has_weight = any('weight' in str(f.get('api_name', '')).lower() for f in fields)
        has_origin = any('origin' in str(f.get('api_name', '')).lower() or '原産' in str(f.get('api_name', '')) or '원산' in str(f.get('api_name', '')) for f in fields)

        all_platforms.append({
            'name': pname.replace('_', ' '), 'hq': pdata.get('hq_country', ''), 'type': ptype,
            'fields': len(fields), 'has_mat': has_mat, 'mat_loc': mat_loc[:50],
            'has_comp': has_comp, 'has_proc': has_proc, 'has_cat': has_cat, 'has_weight': has_weight, 'has_origin': has_origin
        })

for p in all_platforms:
    ws5.cell(row=r, column=1, value=p['name']).font = NF
    ws5.cell(row=r, column=2, value=p['hq']).font = NF
    ws5.cell(row=r, column=3, value=p['type']).font = NF
    ws5.cell(row=r, column=4, value=p['fields']).font = NF
    for ci, key in enumerate(['has_mat', 'has_comp', 'has_proc', 'has_cat', 'has_weight', 'has_origin'], 5):
        if key == 'has_mat':
            # Skip — next column is mat_loc
            pass
    # Material
    cell = ws5.cell(row=r, column=5, value='YES' if p['has_mat'] else 'No')
    cell.fill = GF if p['has_mat'] else RF; cell.font = NF
    ws5.cell(row=r, column=6, value=p['mat_loc']).font = NF
    # Composition, Processing, Category, Weight, Origin
    for ci, val in enumerate([p['has_comp'], p['has_proc'], p['has_cat'], p['has_weight'], p['has_origin']], 7):
        cell = ws5.cell(row=r, column=ci, value='YES' if val else 'No')
        cell.fill = GF if val else RF; cell.font = NF
    for c in range(1, len(h5)+1): ws5.cell(row=r, column=c).border = TB
    r += 1

# Summary stats
r += 1; ws5.cell(row=r, column=1, value='Summary Statistics').font = SF; r += 1
total = len(all_platforms)
mat_yes = sum(1 for p in all_platforms if p['has_mat'])
comp_yes = sum(1 for p in all_platforms if p['has_comp'])
proc_yes = sum(1 for p in all_platforms if p['has_proc'])
origin_yes = sum(1 for p in all_platforms if p['has_origin'])
avg_fields = sum(p['fields'] for p in all_platforms) / total

stats = [
    (f'Total platforms surveyed', total),
    (f'Average product fields per platform', f'{avg_fields:.1f}'),
    (f'Has DEDICATED Material field', f'{mat_yes}/{total} ({mat_yes/total*100:.0f}%)'),
    (f'Has Composition field', f'{comp_yes}/{total} ({comp_yes/total*100:.0f}%)'),
    (f'Has Processing field', f'{proc_yes}/{total} ({proc_yes/total*100:.0f}%)'),
    (f'Has Country of Origin field', f'{origin_yes}/{total} ({origin_yes/total*100:.0f}%)'),
    ('', ''),
    ('Key Findings:', ''),
    ('1. Material as dedicated field', f'{mat_yes}/{total} platforms — mostly Amazon, B2B, Etsy, Zalando, SHEIN, Korean/Chinese platforms'),
    ('2. Composition as structured field', f'{comp_yes}/{total} platforms — Amazon, Zalando, SHEIN, Wildberries (textile-focused)'),
    ('3. Processing as dedicated field', f'{proc_yes}/{total} platforms — only B2B platforms (Alibaba.com Processing Service)'),
    ('4. B2B vs B2C', 'B2B has FAR more detailed material/spec data (alloy grades, tolerances, certifications)'),
    ('5. East Asian platforms (Taobao, Coupang, Naver)', 'Have category-specific required attributes including material — driven by local laws'),
    ('6. Western B2C (Shopify, WooCommerce, BigCommerce)', 'NO built-in material field — all via custom attributes/options'),
]
for lbl, val in stats:
    ws5.cell(row=r, column=1, value=lbl).font = BF if not lbl.startswith(' ') and lbl else NF
    ws5.cell(row=r, column=2, value=str(val)).font = NF
    r += 1

aw(ws5, len(h5), 50)

# Sheet 6: Country × Platform cross-reference
ws6 = wb.create_sheet('Country x Platform')
r = 1; ws6.cell(row=r, column=1, value='Country × Platform — Customs vs E-commerce Field Comparison').font = TF; r = 3
h6 = ['Country', 'Customs Fields (from customs investigation)', 'Platform 1', 'Platform 1 Fields', 'Platform 2', 'Platform 2 Fields', 'Gap Analysis']
for c, h in enumerate(h6, 1): ws6.cell(row=r, column=c, value=h)
hdr(ws6, r, len(h6)); r += 1

cross = [
    ('CN', '18 fields (申报要素: 材质, 成分, 加工方式 etc.)', 'AliExpress', '~10 (material in attributes, no composition)', 'Taobao', '~12 (材质/成分 in props)', 'Gap: AliExpress missing composition/processing that customs requires'),
    ('US', '9 fields (Box 28 free-text)', 'Amazon', '45-200 (category-specific, material dedicated)', 'eBay', '~12 + aspects', 'Amazon has MORE data than customs requires'),
    ('KR', '10 fields (품명, 규격, 원산지 mandatory)', 'Coupang', '~13 (원산지 mandatory, 속성)', 'Naver', '~9 (상품정보제공고시 mandatory)', 'Korean platforms enforce 원산지 + 재질 by law'),
    ('JP', '8 fields (品名, 重量, 原産地)', 'Rakuten', '~11 (simple, no material)', 'Yahoo JP', '~9 (스펙 key-value)', 'Japanese platforms have LESS structured data than customs'),
    ('DE', 'EU SAD 9 fields + textile composition', 'Zalando', '~12 (MOST detailed material_composition in world)', 'Otto', '~9 (material for textiles)', 'Zalando exceeds customs requirements for material data'),
    ('IN', 'BoE ~9 fields + BIS cert', 'Flipkart', '~13 (origin, manufacturer, MRP mandatory)', 'IndiaMART', '~11 (material dedicated, B2B)', 'Indian law forces platforms to collect origin+manufacturer'),
    ('RU', 'ДТ ~10 fields + brand/model', 'Ozon', '~11 (origin mandatory, material in attrs)', 'Wildberries', '~11 (Состав mandatory for clothing)', 'Russian platforms enforce composition for textiles'),
]
for c_data in cross:
    for c, v in enumerate(c_data, 1):
        ws6.cell(row=r, column=c, value=v).font = NF; ws6.cell(row=r, column=c).border = TB; ws6.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
    r += 1
aw(ws6, len(h6), 50)

wb.save(OUTPUT)
print(f'✅ Excel saved: {OUTPUT}')
print(f'   6 sheets, {total} platforms')
