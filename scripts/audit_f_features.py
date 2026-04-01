#!/usr/bin/env python3
"""
F-Feature Audit: Check all 67 command files for execution status.
Outputs POTAL_F_FEATURE_AUDIT_RESULT.xlsx
"""

import os, re, subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = "/Users/maegbug/potal"

# 67 command files in order
CMD_FILES = [
    # Group A: _100 series (6)
    "CLAUDE_CODE_F006_CONFIDENCE_100.md",
    "CLAUDE_CODE_F012_VALIDATION_100.md",
    "CLAUDE_CODE_F046_WEBHOOK_100.md",
    "CLAUDE_CODE_F052_API_AUTH_100.md",
    "CLAUDE_CODE_F093_WEBHOOK_SECURITY_100.md",
    "CLAUDE_CODE_F125_API_KEY_SECURITY_100.md",
    # Group B: _ONLY series (11)
    "CLAUDE_CODE_F013_ONLY.md",
    "CLAUDE_CODE_F015_ONLY.md",
    "CLAUDE_CODE_F026_ONLY.md",
    "CLAUDE_CODE_F041_ONLY.md",
    "CLAUDE_CODE_F049_ONLY.md",
    "CLAUDE_CODE_F054_ONLY.md",
    "CLAUDE_CODE_F068_ONLY.md",
    "CLAUDE_CODE_F081_ONLY.md",
    "CLAUDE_CODE_F090_SDK_ONLY.md",
    "CLAUDE_CODE_F126_RAG_ONLY.md",
    "CLAUDE_CODE_F143_CHATBOT_ONLY.md",
    # Group C: P0 (9)
    "CLAUDE_CODE_F025_DDP_DDU.md",
    "CLAUDE_CODE_F033_IOSS_OSS.md",
    "CLAUDE_CODE_F095_HIGH_THROUGHPUT.md",
    "CLAUDE_CODE_F109_CSV_EXPORT.md",
    "CLAUDE_CODE_F008_AUDIT_TRAIL.md",
    "CLAUDE_CODE_F092_SANDBOX.md",
    "CLAUDE_CODE_F009_BATCH_CLASSIFY.md",
    "CLAUDE_CODE_F043_CUSTOMS_DOCS.md",
    "CLAUDE_CODE_F040_PRE_SHIPMENT.md",
    # Group D: P1 (9)
    "CLAUDE_CODE_F002_IMAGE_CLASSIFY.md",
    "CLAUDE_CODE_F003_URL_CLASSIFY.md",
    "CLAUDE_CODE_F007_ECCN_CLASSIFY.md",
    "CLAUDE_CODE_F037_EXPORT_CONTROLS.md",
    "CLAUDE_CODE_F039_RULES_OF_ORIGIN.md",
    "CLAUDE_CODE_F050_TYPE86.md",
    "CLAUDE_CODE_F097_AI_CONSULT.md",
    "CLAUDE_CODE_F112_WHITELABEL.md",
    "CLAUDE_CODE_F116_MULTILINGUAL.md",
    # Group E: P2 Tax (9)
    "CLAUDE_CODE_F027_US_SALES_TAX.md",
    "CLAUDE_CODE_F028_TELECOM_TAX.md",
    "CLAUDE_CODE_F029_LODGING_TAX.md",
    "CLAUDE_CODE_F038_EXPORT_LICENSE.md",
    "CLAUDE_CODE_F044_CUSTOMS_DECLARATION.md",
    "CLAUDE_CODE_F051_TAX_FILING.md",
    "CLAUDE_CODE_F053_TAX_EXEMPTION.md",
    "CLAUDE_CODE_F055_VAT_REGISTRATION.md",
    "CLAUDE_CODE_F057_EINVOICE.md",
    # Group F: P2 Integration (7)
    "CLAUDE_CODE_F082_MARKETPLACE.md",
    "CLAUDE_CODE_F083_ERP.md",
    "CLAUDE_CODE_F104_TAX_LIABILITY.md",
    "CLAUDE_CODE_F105_COMPLIANCE_AUDIT.md",
    "CLAUDE_CODE_F138_CSM.md",
    "CLAUDE_CODE_F140_AEO.md",
    "CLAUDE_CODE_F147_REVENUE_SHARE.md",
    # Group G: Bundle (16)
    "CLAUDE_CODE_F060_MULTICARRIER.md",
    "CLAUDE_CODE_F061_SHIPPING_LABEL.md",
    "CLAUDE_CODE_F062_TRACKING.md",
    "CLAUDE_CODE_F063_F064_F065_SHIPPING.md",
    "CLAUDE_CODE_F069_F047_F048_CUSTOMS.md",
    "CLAUDE_CODE_F071_F073_F115_CHECKOUT.md",
    "CLAUDE_CODE_F084_ACCOUNTING.md",
    "CLAUDE_CODE_F087_PARTNER_ECOSYSTEM.md",
    "CLAUDE_CODE_F103_F107_ANALYTICS.md",
    "CLAUDE_CODE_F110_F111_BRANDING.md",
    "CLAUDE_CODE_F130_F131_F132_COMMERCE.md",
    "CLAUDE_CODE_F133_F134_ORDERS.md",
    "CLAUDE_CODE_F135_F136_F137_FULFILLMENT.md",
    "CLAUDE_CODE_F141_F144_F145_EDUCATION.md",
    "CLAUDE_CODE_F030_F056_TAX_MISC.md",
    "CLAUDE_CODE_F146_PARTNER_MGMT.md",
]

def extract_f_numbers(filename):
    """Extract F-numbers from filename."""
    matches = re.findall(r'F(\d+)', filename)
    return ['F' + m for m in matches]

def extract_target_files(md_path):
    """Extract .ts/.tsx file paths from command file."""
    targets = []
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except:
        return targets

    # Match patterns like `app/...route.ts`, `app/lib/...ts`
    # Look for backtick-wrapped paths or paths after - or *
    patterns = [
        r'`(app/[^\s`]+\.tsx?)`',
        r'`(sdk/[^\s`]+\.tsx?)`',
        r'`(plugins/[^\s`]+\.tsx?)`',
        r'`(components/[^\s`]+\.tsx?)`',
        r'`(supabase/[^\s`]+\.sql)`',
        r'[-*]\s*(app/[^\s]+\.tsx?)',
        r'[-*]\s*(sdk/[^\s]+\.tsx?)',
        r'파일[:\s]*`?(app/[^\s`]+\.tsx?)`?',
        r'파일[:\s]*`?(sdk/[^\s`]+\.tsx?)`?',
    ]

    for pat in patterns:
        for m in re.finditer(pat, content):
            path = m.group(1).strip().rstrip(')')
            if path not in targets:
                targets.append(path)

    # Also look for route.ts paths mentioned inline
    for m in re.finditer(r'(app/api/v1/[^\s`"\']+/route\.ts)', content):
        path = m.group(1).strip()
        if path not in targets:
            targets.append(path)

    # lib files
    for m in re.finditer(r'(app/lib/[^\s`"\']+\.ts)', content):
        path = m.group(1).strip()
        if path not in targets:
            targets.append(path)

    return targets

def extract_critical_keywords(md_path):
    """Extract key function/feature names from CRITICAL sections."""
    keywords = []
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except:
        return keywords

    # Look for function names in code blocks after CRITICAL
    for m in re.finditer(r'(?:function|async function|export function|export async function)\s+(\w+)', content):
        kw = m.group(1)
        if kw not in keywords and len(kw) > 3:
            keywords.append(kw)

    # Look for const/interface names
    for m in re.finditer(r'(?:const|interface|type)\s+([A-Z_][A-Z_0-9a-z]+)', content):
        kw = m.group(1)
        if kw not in keywords and len(kw) > 5:
            keywords.append(kw)

    return keywords[:10]  # max 10 keywords

def check_file(filepath):
    """Check if file exists, get line count."""
    full = os.path.join(ROOT, filepath)
    if not os.path.exists(full):
        return False, 0
    try:
        with open(full, 'r', encoding='utf-8') as f:
            lines = len(f.readlines())
        return True, lines
    except:
        return False, 0

def check_keywords_in_file(filepath, keywords):
    """Check how many critical keywords are present in the file."""
    full = os.path.join(ROOT, filepath)
    if not os.path.exists(full):
        return 0, 0, []
    try:
        with open(full, 'r', encoding='utf-8') as f:
            content = f.read()
    except:
        return 0, 0, []

    found = [kw for kw in keywords if kw in content]
    return len(found), len(keywords), found

def judge(exists, lines, kw_found, kw_total):
    """Final judgment."""
    if not exists:
        return "NOT_EXECUTED", "File not found"
    if lines < 20:
        return "NOT_EXECUTED", f"Only {lines} lines (empty/minimal)"
    if kw_total == 0:
        # No keywords extracted from command file — judge by line count
        if lines >= 50:
            return "EXECUTED", f"{lines} lines, no keywords to verify"
        else:
            return "PARTIAL", f"{lines} lines, no keywords to verify"
    ratio = kw_found / kw_total if kw_total > 0 else 0
    if ratio >= 0.5:
        return "EXECUTED", f"{kw_found}/{kw_total} keywords found"
    elif ratio > 0 or lines >= 50:
        return "PARTIAL", f"{kw_found}/{kw_total} keywords, {lines} lines"
    else:
        return "NOT_EXECUTED", f"{kw_found}/{kw_total} keywords"


# ─── Main ────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Audit Result"

# Headers
headers = ["#", "Command File", "F Numbers", "Target File", "Exists", "Lines", "Critical Keywords", "Judgment", "Notes"]
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 7
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 40

row = 2
executed = 0
partial = 0
not_executed = 0
cmd_judgments = {}  # per command file: best judgment

green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

for idx, cmd_file in enumerate(CMD_FILES, 1):
    md_path = os.path.join(ROOT, cmd_file)
    f_numbers = extract_f_numbers(cmd_file)
    f_str = ", ".join(f_numbers)

    targets = extract_target_files(md_path)
    keywords = extract_critical_keywords(md_path)

    if not targets:
        # No target files found in command file
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=cmd_file).border = thin_border
        ws.cell(row=row, column=3, value=f_str).border = thin_border
        ws.cell(row=row, column=4, value="(no target files extracted)").border = thin_border
        ws.cell(row=row, column=5, value="-").border = thin_border
        ws.cell(row=row, column=6, value=0).border = thin_border
        ws.cell(row=row, column=7, value="-").border = thin_border
        c = ws.cell(row=row, column=8, value="UNKNOWN")
        c.border = thin_border
        c.fill = yellow_fill
        ws.cell(row=row, column=9, value="Could not extract target files from command file").border = thin_border
        row += 1
        continue

    best_judgment = "NOT_EXECUTED"

    for target in targets:
        exists, lines = check_file(target)
        kw_found, kw_total, found_list = check_keywords_in_file(target, keywords)
        judgment, notes = judge(exists, lines, kw_found, kw_total)

        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=cmd_file).border = thin_border
        ws.cell(row=row, column=3, value=f_str).border = thin_border
        ws.cell(row=row, column=4, value=target).border = thin_border
        ws.cell(row=row, column=5, value="Y" if exists else "N").border = thin_border
        ws.cell(row=row, column=6, value=lines).border = thin_border
        ws.cell(row=row, column=7, value=f"{kw_found}/{kw_total}" if kw_total > 0 else "-").border = thin_border

        c = ws.cell(row=row, column=8, value=judgment)
        c.border = thin_border
        if judgment == "EXECUTED":
            c.fill = green_fill
        elif judgment == "PARTIAL":
            c.fill = yellow_fill
        else:
            c.fill = red_fill

        ws.cell(row=row, column=9, value=notes).border = thin_border

        # Track best per command file
        if judgment == "EXECUTED":
            best_judgment = "EXECUTED"
        elif judgment == "PARTIAL" and best_judgment != "EXECUTED":
            best_judgment = "PARTIAL"

        row += 1

    cmd_judgments[cmd_file] = best_judgment

# Summary
row += 1
for cmd, j in cmd_judgments.items():
    if j == "EXECUTED":
        executed += 1
    elif j == "PARTIAL":
        partial += 1
    else:
        not_executed += 1

ws.cell(row=row, column=1, value="").border = thin_border
ws.cell(row=row, column=2, value="=== SUMMARY ===").font = Font(bold=True, size=12)
row += 1
ws.cell(row=row, column=2, value=f"Total Command Files: {len(CMD_FILES)}")
row += 1
c = ws.cell(row=row, column=2, value=f"✅ EXECUTED: {executed}")
c.fill = green_fill
c.font = Font(bold=True)
row += 1
c = ws.cell(row=row, column=2, value=f"⚠️ PARTIAL: {partial}")
c.fill = yellow_fill
c.font = Font(bold=True)
row += 1
c = ws.cell(row=row, column=2, value=f"❌ NOT_EXECUTED: {not_executed}")
c.fill = red_fill
c.font = Font(bold=True)

out_path = os.path.join(ROOT, "POTAL_F_FEATURE_AUDIT_RESULT.xlsx")
wb.save(out_path)
print(f"Saved to {out_path}")
print(f"EXECUTED={executed} PARTIAL={partial} NOT_EXECUTED={not_executed} TOTAL={len(CMD_FILES)}")
